# app.py
# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import io
import json
import re
import secrets
import time
from dataclasses import dataclass
from urllib.parse import urlencode
from datetime import datetime, date, timedelta
from typing import Any, Dict, List, Optional, Tuple, Callable

import requests
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


# =========================
# Config
# =========================
FREEE_TOKEN_URL = "https://accounts.secure.freee.co.jp/public_api/token"
FREEE_AUTHORIZE_URL = "https://accounts.secure.freee.co.jp/public_api/authorize"
FREEE_HR_API_BASE = "https://api.freee.co.jp/hr/api/v1"
TOKEN_STORE_PATH = "token_store.json"
OAUTH_STATE_PATH = ".oauth_state"
DOTENV_PATH = ".env"

COMPANY_ID = 12018586  # 固定

# --- 高速化の要点 ---
# True: 実績がある日は原則 work_records PUT で反映（list_time_clocks を呼ばない）
#       → 最速。time_clocks(打刻履歴)を作らない可能性がある（ただし勤怠画面の結果は同じになる想定）
# False: 従来通り、打刻があるか確認して、無ければ time_clocks POST（厳密モード）
FAST_SYNC_PUT_ONLY = True

# Jobcan summary layout
ROW_STAFF_CODE = 4
COL_STAFF_CODE = 3       # C4
ROW_DAILY_HEADER = 10

COL_DATE = 1
COL_KINTAI_STATUS = 2
COL_HOLIDAY = 3
COL_SHIFT_START = 6
COL_SHIFT_END = 7
COL_CLOCK_IN = 8
COL_CLOCK_OUT = 9
COL_BREAK = 14
COL_PAID_LEAVE = 15

VALID_TIME_CLOCK_TYPES = ["clock_in", "break_begin", "break_end", "clock_out"]

SCHEDULE_COLUMNS_FIXED = [
    "従業員番号",
    "freee人事労務での表示名（編集しても反映されません）",
    "日付",
    "勤務パターンコード",
    "勤務日種別",
    "出勤時刻",
    "退勤時刻",
    "休憩時間",
    "休憩開始1",
    "休憩終了1",
    "休憩開始2",
    "休憩終了2",
    "休憩開始3",
    "休憩終了3",
    "夜勤日種別",
]


# =========================
# Utils
# =========================
def safe_str(v: Any) -> str:
    return "" if v is None else str(v).strip()

def jst_now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")

def load_dotenv_simple(path: str = DOTENV_PATH, override: bool = False) -> None:
    if not os.path.exists(path):
        return
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            s = line.strip()
            if not s or s.startswith("#") or "=" not in s:
                continue
            k, v = s.split("=", 1)
            k = k.strip()
            v = v.strip()
            if v.startswith("#"):
                v = ""
            if (len(v) >= 2) and ((v[0] == v[-1]) and v[0] in ("'", '"')):
                v = v[1:-1]
            if not override and k in os.environ and os.environ[k].strip() != "":
                continue
            os.environ[k] = v

def normalize_employee_number(s: Any) -> str:
    t = safe_str(s)
    if t == "":
        return ""
    trans = str.maketrans("０１２３４５６７８９", "0123456789")
    t = t.translate(trans).strip()
    t2 = t.lstrip("0")
    return t2 if t2 != "" else "0"

def parse_hhmm(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if hasattr(value, "hour") and hasattr(value, "minute"):
        return f"{int(value.hour):02d}:{int(value.minute):02d}"

    s = str(value).strip()
    if s == "":
        return None

    m = re.match(r"^(\d{1,2}):(\d{1,2})$", s)
    if m:
        h = int(m.group(1)); mm = int(m.group(2))
        return f"{h:02d}:{mm:02d}"

    m = re.match(r"^(\d{2})(\d{2})$", s)
    if m:
        return f"{int(m.group(1)):02d}:{int(m.group(2)):02d}"

    raise ValueError(f"時刻の解釈に失敗: {value!r}")

def parse_yyyymm_from_a1(v: Any) -> Optional[Tuple[int, int]]:
    s = safe_str(v)
    m = re.match(r"^\s*(\d{4})年\s*(\d{1,2})月\s*$", s)
    if not m:
        return None
    return int(m.group(1)), int(m.group(2))

def parse_mmdd_youbi(s: str, year: int) -> date:
    m = re.match(r"^\s*(\d{1,2})/(\d{1,2})", s)
    if not m:
        raise ValueError(f"日付文字列の解釈に失敗: {s!r}")
    month = int(m.group(1)); day = int(m.group(2))
    return date(year, month, day)

def combine_date_time(d: date, hhmm: str) -> datetime:
    h, m = map(int, hhmm.split(":"))
    day_add = 0
    if h >= 24:
        day_add = h // 24
        h = h % 24
    base = datetime(d.year, d.month, d.day, h, m, 0)
    if day_add:
        base += timedelta(days=day_add)
    return base

def break_hhmm_to_minutes(hhmm: Optional[str]) -> int:
    if not hhmm:
        return 0
    h, m = map(int, hhmm.split(":"))
    return h * 60 + m

def minutes_to_hhmm(mins: int) -> str:
    if mins <= 0:
        return ""
    h = mins // 60
    m = mins % 60
    return f"{h:02d}:{m:02d}"

def parse_duration_to_minutes(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, datetime):
        return int(value.hour) * 60 + int(value.minute)
    if hasattr(value, "hour") and hasattr(value, "minute"):
        return int(value.hour) * 60 + int(value.minute)
    if isinstance(value, (int, float)):
        if value <= 0:
            return 0
        return int(round(float(value) * 60))
    s = str(value).strip()
    if s == "":
        return 0
    m = re.match(r"^(\d{1,2}):(\d{1,2})$", s)
    if m:
        hh = int(m.group(1)); mm = int(m.group(2))
        if hh < 0 or mm < 0:
            return 0
        return hh * 60 + mm
    try:
        f = float(s)
        if f <= 0:
            return 0
        return int(round(f * 60))
    except Exception:
        return 0

def compute_break_segment(start_dt: datetime, end_dt: datetime, break_mins: int) -> Optional[Tuple[datetime, datetime]]:
    if break_mins <= 0:
        return None
    total = (end_dt - start_dt).total_seconds() / 60
    if total <= break_mins + 1:
        return None
    mid = start_dt + timedelta(minutes=total / 2)
    b_start = (mid - timedelta(minutes=break_mins / 2)).replace(second=0, microsecond=0)
    b_end = (b_start + timedelta(minutes=break_mins)).replace(second=0, microsecond=0)
    return b_start, b_end

def compute_shift_duration_minutes(wd: date, shift_start: Optional[str], shift_end: Optional[str]) -> int:
    if not shift_start or not shift_end:
        return 0
    sdt = combine_date_time(wd, shift_start)
    edt = combine_date_time(wd, shift_end)
    if edt <= sdt:
        edt += timedelta(days=1)
    mins = int(round((edt - sdt).total_seconds() / 60))
    return max(0, mins)

def df_to_csv_bytes_utf8sig(df: pd.DataFrame) -> bytes:
    buf = io.StringIO()
    df.to_csv(buf, index=False)
    return buf.getvalue().encode("utf-8-sig")

def normalize_holiday_kbn(raw: Any) -> str:
    s = safe_str(raw)
    if s == "":
        return "所定労働日"
    s2 = s.replace("　", " ").replace(",", ", ").strip()
    if "法休" in s2:
        return "法定休日"
    if "公休" in s2:
        return "所定休日"
    return "所定労働日"

def is_zero_zero_shift(shift_start: Optional[str], shift_end: Optional[str]) -> bool:
    s = (shift_start or "").strip()
    e = (shift_end or "").strip()
    if s == "" and e == "":
        return True
    return (s in ("0:00", "00:00") and e in ("0:00", "00:00"))

def clamp_int(v: int, lo: int, hi: int) -> int:
    return max(lo, min(hi, v))


# =========================
# Cross-midnight detection
# =========================
def normalize_day_segments(
    wd: date, clock_in: str, clock_out: str, break_mins: int
) -> Tuple[datetime, datetime, Optional[Tuple[datetime, datetime]]]:
    sdt = combine_date_time(wd, clock_in)
    edt = combine_date_time(wd, clock_out)
    if edt <= sdt:
        edt += timedelta(days=1)
    bseg = compute_break_segment(sdt, edt, break_mins)
    return sdt, edt, bseg

def find_cross_midnight_before_statutory_holiday(
    employee_map: Dict[str, Tuple[int, str]],
    staff_rows: Dict[str, List[Dict[str, Any]]],
) -> pd.DataFrame:
    rows_out: List[Dict[str, Any]] = []

    for staff_code, rows in staff_rows.items():
        emp_name = employee_map.get(staff_code, (None, ""))[1]
        by_date: Dict[str, Dict[str, Any]] = {r["work_date"].strftime("%Y-%m-%d"): r for r in rows}

        for r in rows:
            wd: date = r["work_date"]
            day_type = normalize_holiday_kbn(r.get("holiday_kbn"))
            if day_type != "法定休日":
                continue

            prev_date = wd - timedelta(days=1)
            prev_key = prev_date.strftime("%Y-%m-%d")
            prev = by_date.get(prev_key)
            if not prev:
                continue
            if not prev.get("clock_in") or not prev.get("clock_out"):
                continue

            try:
                sdt, edt, _ = normalize_day_segments(
                    prev_date,
                    prev["clock_in"],
                    prev["clock_out"],
                    int(prev.get("break_mins", 0) or 0),
                )
            except Exception as e:
                rows_out.append({
                    "employee_number": staff_code,
                    "employee_name": emp_name,
                    "statutory_holiday_date": wd.strftime("%Y-%m-%d"),
                    "prev_date": prev_key,
                    "prev_clock_in": safe_str(prev.get("clock_in")),
                    "prev_clock_out": safe_str(prev.get("clock_out")),
                    "crossed_to": "",
                    "note": f"parse_error: {e}",
                })
                continue

            crossed = (edt.date() != sdt.date())
            if crossed and edt.date() == wd:
                rows_out.append({
                    "employee_number": staff_code,
                    "employee_name": emp_name,
                    "statutory_holiday_date": wd.strftime("%Y-%m-%d"),
                    "prev_date": prev_key,
                    "prev_clock_in": sdt.strftime("%Y-%m-%d %H:%M"),
                    "prev_clock_out": edt.strftime("%Y-%m-%d %H:%M"),
                    "crossed_to": wd.strftime("%Y-%m-%d"),
                    "note": "prev_day_clock_out_cross_midnight_to_statutory_holiday",
                })

    df = pd.DataFrame(rows_out)
    if not df.empty:
        df = df.sort_values(["employee_number", "prev_date"]).reset_index(drop=True)
    return df


# =========================
# Token Store
# =========================
@dataclass
class FreeeToken:
    access_token: str
    refresh_token: str
    expires_at: Optional[float] = None

    @staticmethod
    def load() -> Optional["FreeeToken"]:
        if os.path.exists(TOKEN_STORE_PATH):
            with open(TOKEN_STORE_PATH, "r", encoding="utf-8") as f:
                d = json.load(f)
            return FreeeToken(
                access_token=d.get("access_token", ""),
                refresh_token=d.get("refresh_token", ""),
                expires_at=d.get("expires_at"),
            )
        return None

    def save(self) -> None:
        with open(TOKEN_STORE_PATH, "w", encoding="utf-8") as f:
            json.dump(
                {
                    "access_token": self.access_token,
                    "refresh_token": self.refresh_token,
                    "expires_at": self.expires_at,
                    "saved_at": jst_now_iso(),
                },
                f,
                ensure_ascii=False,
                indent=2,
            )


def get_redirect_uri() -> str:
    """.env の FREEE_REDIRECT_URI、未設定時は http://localhost:8080/callback"""
    uri = os.environ.get("FREEE_REDIRECT_URI", "").strip()
    if uri:
        return uri.rstrip("/") if uri.endswith("/") and uri != "http://" else uri
    return "http://localhost:8080/callback"


def exchange_code_for_token(
    client_id: str, client_secret: str, code: str, redirect_uri: str
) -> FreeeToken:
    """認可コードをアクセス・リフレッシュトークンに交換し、token_store.json に保存する。"""
    resp = requests.post(
        FREEE_TOKEN_URL,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        data={
            "grant_type": "authorization_code",
            "client_id": client_id,
            "client_secret": client_secret,
            "code": code,
            "redirect_uri": redirect_uri,
        },
        timeout=30,
    )
    if resp.status_code != 200:
        raise RuntimeError(f"トークン取得失敗: {resp.status_code} {resp.text}")
    j = resp.json()
    token = FreeeToken(
        access_token=j["access_token"],
        refresh_token=j.get("refresh_token", ""),
        expires_at=None,
    )
    if j.get("expires_in"):
        token.expires_at = time.time() + float(j["expires_in"]) - 60
    token.save()
    return token


def build_authorize_url(client_id: str, redirect_uri: str, state: str) -> str:
    """freee 認可画面のURLを組み立てる。"""
    params = {
        "response_type": "code",
        "client_id": client_id,
        "redirect_uri": redirect_uri,
        "state": state,
        "prompt": "select_company",
    }
    return f"{FREEE_AUTHORIZE_URL}?{urlencode(params)}"


# =========================
# freee HR Client (高速版: Session + Retry + Pool)
# =========================
class FreeeHrClient:
    def __init__(self, client_id: str, client_secret: str, token: FreeeToken):
        self.client_id = str(client_id)
        self.client_secret = str(client_secret)
        self.token = token

        self.session = requests.Session()
        retry = Retry(
            total=3,
            backoff_factor=0.6,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["GET", "POST", "PUT"],
            raise_on_status=False,
        )
        adapter = HTTPAdapter(max_retries=retry, pool_connections=20, pool_maxsize=20)
        self.session.mount("https://", adapter)
        self.session.mount("http://", adapter)

    def _refresh(self) -> None:
        def do_post(refresh_token_value: str) -> requests.Response:
            return requests.post(
                FREEE_TOKEN_URL,
                headers={"Content-Type": "application/x-www-form-urlencoded"},
                data={
                    "grant_type": "refresh_token",
                    "client_id": self.client_id,
                    "client_secret": self.client_secret,
                    "refresh_token": refresh_token_value,
                },
                timeout=30,
            )

        resp = do_post(self.token.refresh_token)
        if resp.status_code != 200:
            # token_store のトークンが期限切れのとき、.env の FREEE_REFRESH_TOKEN で1回だけ再試行
            env_refresh = os.environ.get("FREEE_REFRESH_TOKEN", "").strip()
            if env_refresh and env_refresh != self.token.refresh_token:
                self.token.refresh_token = env_refresh
                resp = do_post(self.token.refresh_token)
            if resp.status_code != 200:
                raise RuntimeError(f"Refresh failed: {resp.status_code} {resp.text}")
        j = resp.json()
        self.token.access_token = j["access_token"]
        self.token.refresh_token = j.get("refresh_token", self.token.refresh_token)
        expires_in = j.get("expires_in")
        if expires_in:
            self.token.expires_at = time.time() + float(expires_in) - 60
        self.token.save()

    def _maybe_refresh_before_request(self) -> None:
        if self.token.expires_at and time.time() > self.token.expires_at:
            self._refresh()

    def request(self, method: str, path: str, *, params=None, json_body=None) -> requests.Response:
        self._maybe_refresh_before_request()
        url = f"{FREEE_HR_API_BASE}{path}"
        headers = {"Authorization": f"Bearer {self.token.access_token}"}

        resp = self.session.request(method, url, headers=headers, params=params, json=json_body, timeout=30)
        if resp.status_code == 401:
            self._refresh()
            headers = {"Authorization": f"Bearer {self.token.access_token}"}
            resp = self.session.request(method, url, headers=headers, params=params, json=json_body, timeout=30)
        if resp.status_code == 403:
            try:
                msg = resp.json().get("message", "")
            except Exception:
                msg = resp.text
            if "アクセス権限" in msg or "scope" in msg.lower():
                raise RuntimeError(
                    "freee APIから403エラー（権限不足）: アプリのスコープが更新されています。"
                    "サイドバーの「認証を開始」からfreeeの再認可を行ってください。\n"
                    f"詳細: {msg}"
                )
        return resp

    def get_company_employees_all(self, company_id: int, limit: int = 100, year: Optional[int] = None, month: Optional[int] = None) -> List[Dict[str, Any]]:
        all_items: List[Dict[str, Any]] = []
        offset = 0
        while True:
            params: Dict[str, Any] = {"limit": limit, "offset": offset}
            if year is not None:
                params["year"] = year
            if month is not None:
                params["month"] = month
            r = self.request("GET", f"/companies/{company_id}/employees", params=params)
            r.raise_for_status()
            payload = r.json()

            if isinstance(payload, list):
                items = payload
            elif isinstance(payload, dict):
                for k in ["employees", "items", "data"]:
                    if k in payload and isinstance(payload[k], list):
                        items = payload[k]
                        break
                else:
                    items = [payload] if "id" in payload else []
            else:
                items = []

            all_items.extend(items)
            if len(items) < limit:
                break
            offset += limit
        return all_items

    def list_time_clocks(self, employee_id: int, company_id: int, base_date: str) -> requests.Response:
        return self.request("GET", f"/employees/{employee_id}/time_clocks", params={"company_id": company_id, "base_date": base_date})

    def delete_time_clock(self, employee_id: int, time_clock_id: int, company_id: int) -> requests.Response:
        return self.request("DELETE", f"/employees/{employee_id}/time_clocks/{time_clock_id}", params={"company_id": company_id})

    def post_time_clock(self, employee_id: int, company_id: int, ttype: str, base_date: str, dt_str: str) -> requests.Response:
        body = {"company_id": company_id, "type": ttype, "base_date": base_date, "datetime": dt_str}
        return self.request("POST", f"/employees/{employee_id}/time_clocks", json_body=body)

    def put_work_record(self, employee_id: int, target_date: str, body: Dict[str, Any]) -> requests.Response:
        return self.request("PUT", f"/employees/{employee_id}/work_records/{target_date}", json_body=body)


def build_employee_map(employees_list: List[Dict[str, Any]]) -> Dict[str, Tuple[int, str]]:
    number_keys = ["employee_number", "num", "code", "external_id"]
    name_keys = ["display_name", "name", "full_name"]

    out: Dict[str, Tuple[int, str]] = {}
    for e in employees_list:
        if not isinstance(e, dict) or "id" not in e:
            continue

        num = ""
        for k in number_keys:
            if k in e and safe_str(e[k]) != "":
                num = safe_str(e[k])
                break
        if num == "":
            continue

        emp_id = int(e["id"])

        nm = ""
        for nk in name_keys:
            if nk in e and safe_str(e[nk]) != "":
                nm = safe_str(e[nk])
                break
        if nm == "":
            nm = f"employee_id={emp_id}"

        out[normalize_employee_number(num)] = (emp_id, nm)

    return out


# =========================
# Jobcan reader
# =========================
def find_sheet_staff_code(ws) -> str:
    return normalize_employee_number(ws.cell(ROW_STAFF_CODE, COL_STAFF_CODE).value)

def read_one_sheet_rows(ws, year: int, month: int) -> List[Dict[str, Any]]:
    if safe_str(ws.cell(ROW_DAILY_HEADER, COL_DATE).value) != "日付":
        return []

    rows: List[Dict[str, Any]] = []
    r = ROW_DAILY_HEADER + 1

    while r <= ws.max_row:
        s = safe_str(ws.cell(r, COL_DATE).value)
        if s == "" or s in ("合計", "小計", "総計"):
            break
        if not re.match(r"^\d{1,2}/\d{1,2}", s):
            break

        wd = parse_mmdd_youbi(s, year)
        if wd.month != month:
            r += 1
            continue

        kintai_status = safe_str(ws.cell(r, COL_KINTAI_STATUS).value)
        holiday_kbn = safe_str(ws.cell(r, COL_HOLIDAY).value)

        shift_start = parse_hhmm(ws.cell(r, COL_SHIFT_START).value)
        shift_end = parse_hhmm(ws.cell(r, COL_SHIFT_END).value)
        if shift_start == "00:00" and shift_end == "00:00":
            shift_start, shift_end = None, None

        clock_in = parse_hhmm(ws.cell(r, COL_CLOCK_IN).value)
        clock_out = parse_hhmm(ws.cell(r, COL_CLOCK_OUT).value)
        if clock_in == "00:00" and clock_out == "00:00":
            clock_in, clock_out = None, None

        break_hhmm = parse_hhmm(ws.cell(r, COL_BREAK).value)
        break_mins = break_hhmm_to_minutes(break_hhmm)

        paid_leave_raw = ws.cell(r, COL_PAID_LEAVE).value
        paid_leave_mins = parse_duration_to_minutes(paid_leave_raw)

        rows.append({
            "work_date": wd,
            "kintai_status": kintai_status,
            "holiday_kbn": holiday_kbn,
            "shift_start": shift_start,
            "shift_end": shift_end,
            "clock_in": clock_in,
            "clock_out": clock_out,
            "break_hhmm": break_hhmm,
            "break_mins": break_mins,
            "paid_leave_raw": paid_leave_raw,
            "paid_leave_mins": paid_leave_mins,
            "source_row": r,
        })
        r += 1

    return rows

def read_jobcan_excels(files: List[Tuple[str, bytes]]) -> Tuple[int, int, Dict[str, List[Dict[str, Any]]]]:
    staff_rows: Dict[str, List[Dict[str, Any]]] = {}
    detected_year: Optional[int] = None
    detected_month: Optional[int] = None
    detected_from: Optional[str] = None

    for fname, content in files:
        wb = load_workbook(io.BytesIO(content), data_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]

            ym = parse_yyyymm_from_a1(ws.cell(1, 1).value)
            if ym is None:
                continue
            y, m = ym

            if detected_year is None:
                detected_year, detected_month = y, m
                detected_from = f"{fname}:{sname}"
            else:
                if y != detected_year or m != detected_month:
                    raise ValueError(
                        f"Excelの年月が混在しています。最初={detected_year}年{detected_month}月（{detected_from}） / "
                        f"別={y}年{m}月（{fname}:{sname}）"
                    )

            staff_code = find_sheet_staff_code(ws)
            if staff_code == "":
                continue

            rows = read_one_sheet_rows(ws, detected_year, detected_month)
            if rows:
                staff_rows.setdefault(staff_code, []).extend(rows)

    if detected_year is None or detected_month is None:
        raise ValueError("年月を確定できませんでした。各シートのA1が 'YYYY年MM月' になっているか確認してください。")

    normalized: Dict[str, List[Dict[str, Any]]] = {}
    for code, rows in staff_rows.items():
        by_date: Dict[str, Dict[str, Any]] = {}
        for rec in rows:
            k = rec["work_date"].strftime("%Y-%m-%d")
            by_date[k] = rec
        normalized[code] = [by_date[k] for k in sorted(by_date.keys())]

    return detected_year, detected_month, normalized


# =========================
# Schedule CSV (fixed columns)
# =========================
def build_schedule_csv_fixed(
    employee_map: Dict[str, Tuple[int, str]],
    staff_rows: Dict[str, List[Dict[str, Any]]],
) -> Tuple[pd.DataFrame, bytes]:
    """
    固定列CSVを生成する。

    休憩時間のルール（優先順）:
      1) シフト開始・終了が 0:00/0:00（または空） -> 休憩は必ず空欄
      2) シフト時間が 6時間未満 -> 休憩なし（空欄）
      3) シフト時間が 8時間(480分) -> 休憩は "01:00" 固定
      4) シフト時間が 8時間(480分)を超える -> (シフト - 480分) が休憩
         例) 7:00-17:00(600分) -> 休憩120分(02:00)
      5) それ以外（360〜480の間） -> Excel由来の休憩時間（分）を採用（上限丸め）

    有休行（B列=有休）:
      - 勤務日種別は "所定労働日"
      - 出勤/退勤はシフト開始/終了
      - 休憩は「通常のシフト休憩ルール」で決める（12/8問題対策）
    """
    cols = SCHEDULE_COLUMNS_FIXED[:]
    out_rows: List[Dict[str, Any]] = []

    def new_row() -> Dict[str, Any]:
        return {c: "" for c in cols}

    def calc_break_hhmm_for_shift(
        zero_shift: bool,
        wd: date,
        shift_start: Optional[str],
        shift_end: Optional[str],
        excel_break_mins: int,
    ) -> str:
        shift_mins = compute_shift_duration_minutes(wd, shift_start, shift_end)

        if zero_shift or shift_mins <= 0:
            return ""

        if shift_mins < 360:
            return ""

        if shift_mins == 540:
            return "01:00"

        if shift_mins > 540:
            desired_break = shift_mins - 480
            if desired_break >= shift_mins:
                return ""
            return minutes_to_hhmm(desired_break)

        if excel_break_mins <= 0:
            return ""
        eff = min(excel_break_mins, shift_mins)
        if eff >= shift_mins:
            return ""
        return minutes_to_hhmm(eff)

    for staff_code, rows in staff_rows.items():
        if staff_code not in employee_map:
            continue
        _emp_id, emp_name = employee_map[staff_code]

        for rec in rows:
            wd: date = rec["work_date"]

            holiday_based = normalize_holiday_kbn(rec.get("holiday_kbn"))
            is_paid_leave = (safe_str(rec.get("kintai_status")) == "有休")
            zero_shift = is_zero_zero_shift(rec.get("shift_start"), rec.get("shift_end"))

            row = new_row()
            row["従業員番号"] = staff_code
            row["freee人事労務での表示名（編集しても反映されません）"] = emp_name
            row["日付"] = wd.strftime("%Y-%m-%d")
            row["勤務パターンコード"] = ""
            row["夜勤日種別"] = ""

            if is_paid_leave:
                row["勤務日種別"] = "所定労働日"
            else:
                row["勤務日種別"] = holiday_based

            if is_paid_leave:
                row["出勤時刻"] = rec.get("shift_start") or ""
                row["退勤時刻"] = rec.get("shift_end") or ""
                row["休憩時間"] = calc_break_hhmm_for_shift(
                    zero_shift,
                    wd,
                    rec.get("shift_start"),
                    rec.get("shift_end"),
                    int(rec.get("break_mins", 0) or 0),
                )
            else:
                if row["勤務日種別"] == "所定労働日":
                    row["出勤時刻"] = rec.get("shift_start") or ""
                    row["退勤時刻"] = rec.get("shift_end") or ""
                    row["休憩時間"] = calc_break_hhmm_for_shift(
                        zero_shift,
                        wd,
                        rec.get("shift_start"),
                        rec.get("shift_end"),
                        int(rec.get("break_mins", 0) or 0),
                    )
                else:
                    row["出勤時刻"] = ""
                    row["退勤時刻"] = ""
                    row["休憩時間"] = ""

            out_rows.append(row)

    out = pd.DataFrame(out_rows, columns=cols)
    csv_bytes = df_to_csv_bytes_utf8sig(out)
    return out, csv_bytes


# =========================
# work_records bodies
# =========================

def build_work_records_body(
    company_id: int, target_date: str, sdt: datetime, edt: datetime, bseg: Optional[Tuple[datetime, datetime]]
) -> Dict[str, Any]:
    def hhmm(dt: datetime) -> str:
        # 日跨ぎ対応: 翌日以降は 24:xx, 25:xx 形式
        base = date.fromisoformat(target_date)
        day_diff = (dt.date() - base).days
        h = dt.hour + day_diff * 24
        return f"{h:02d}:{dt.minute:02d}"

    # 休憩がある場合は2セグメントに分割（新API仕様: work_records配列）
    if bseg:
        b_start, b_end = bseg
        segments = [
            {"clock_in_at": hhmm(sdt),     "clock_out_at": hhmm(b_start)},
            {"clock_in_at": hhmm(b_end),    "clock_out_at": hhmm(edt)},
        ]
    else:
        segments = [{"clock_in_at": hhmm(sdt), "clock_out_at": hhmm(edt)}]

    return {
        "company_id": company_id,
        "target_date": target_date,
        "not_auto_calc_work_time": True,
        "work_records": segments,
    }

def build_paid_holiday_work_record_body(company_id: int, target_date: str) -> Dict[str, Any]:
    return {
        "company_id": company_id,
        "target_date": target_date,
        "paid_holiday": 1.0,
        "paid_holidays": [{"type": "full", "days": 1, "mins": 0}],
        "comment": "Jobcan連携: paid holiday by script",
    }

def build_absence_work_record_body(
    company_id: int,
    target_date: str,
    *,
    use_attendance_deduction: bool = True,
    note: str = "Jobcan連携: absence by script",
) -> Dict[str, Any]:
    """
    欠勤登録（work_records PUT）
    - is_absence: 欠勤フラグ
    - clear_work_time: 既存の出退勤・休憩等を消す意図
    - work_records/break_records は空（=勤務実績なし）
    """
    return {
        "company_id": company_id,
        "target_date": target_date,
        "is_absence": True,
        "use_attendance_deduction": bool(use_attendance_deduction),
        "clear_work_time": True,
        "work_records": [],
        "break_records": [],
        "note": note,
        "comment": note,
    }


# =========================
# Sync (高速化版) + 進捗コールバック
# =========================
ProgressCb = Optional[Callable[[int, int, str, str, str], None]]
# progress_cb(done, total, staff_code, emp_name, base_date)

def sync_time_clocks_and_work_records(
    client: "FreeeHrClient",
    employee_map: Dict[str, Tuple[int, str]],
    staff_rows: Dict[str, List[Dict[str, Any]]],
    company_id: int,
    fast_put_only: bool = FAST_SYNC_PUT_ONLY,
    progress_cb: ProgressCb = None,
    *,
    # 欠勤運用の調整用（必要ならUIから渡す）
    absence_only_on_workday: bool = True,           # 所定労働日のみ欠勤PUTする
    absence_use_attendance_deduction: bool = True,  # 欠勤控除対象に算入する
) -> Tuple[pd.DataFrame, bytes]:
    """
    高速化方針:
      fast_put_only=True:
        - 有休: work_records PUT
        - 欠勤: work_records PUT（is_absence）
        - 実績がある日(所定/休日出勤): work_records PUT
        - list_time_clocks / time_clocks POST を行わない（呼び出し回数を最小化）
      fast_put_only=False:
        - 従来通り: list_time_clocks を見て、既存ありは PUT、無しは POST を行う

    進捗表示:
      progress_cb(done, total, staff_code, emp_name, base_date) を 1レコード単位で呼ぶ
    """
    results: List[Dict[str, Any]] = []

    def _log(
        staff_code: str,
        emp_name: str,
        base_date: str,
        action: str,
        resp: Optional[requests.Response],
        body_any: Any,
    ):
        if resp is None:
            results.append({
                "employee_number": staff_code,
                "employee_name": emp_name,
                "date": base_date,
                "action": action,
                "status": "",
                "body": body_any if isinstance(body_any, str) else json.dumps(body_any, ensure_ascii=False),
            })
            return

        try:
            jb = resp.json()
        except Exception:
            jb = resp.text

        results.append({
            "employee_number": staff_code,
            "employee_name": emp_name,
            "date": base_date,
            "action": action,
            "status": resp.status_code,
            "body": jb if isinstance(jb, str) else json.dumps(jb, ensure_ascii=False),
        })

    total = sum(len(v) for v in staff_rows.values())
    done = 0

    for staff_code, rows in staff_rows.items():
        if staff_code not in employee_map:
            # 進捗は「入力レコード」基準で進める（スキップでも進む）
            for rec in rows:
                done += 1
                wd: date = rec["work_date"]
                base_date = wd.strftime("%Y-%m-%d")
                if progress_cb:
                    progress_cb(done, total, staff_code, "", base_date)
            _log(staff_code, "", "", "SKIP", None, "employee not found in freee")
            continue

        employee_id, emp_name = employee_map[staff_code]

        for rec in rows:
            wd: date = rec["work_date"]
            base_date = wd.strftime("%Y-%m-%d")

            done += 1
            if progress_cb:
                progress_cb(done, total, staff_code, emp_name, base_date)

            holiday_based = normalize_holiday_kbn(rec.get("holiday_kbn"))
            status = safe_str(rec.get("kintai_status"))

            is_paid_leave = (status == "有休")
            is_absence = (status == "欠勤")

            # ---- 実績有無（欠勤/有休以外の通常日の判定に使う）----
            has_actual = bool(rec.get("clock_in")) and bool(rec.get("clock_out"))

            # --- 有休は PUT（paid_holiday） ---
            if is_paid_leave:
                body = build_paid_holiday_work_record_body(company_id, base_date)
                resp = client.put_work_record(employee_id, base_date, body)
                _log(staff_code, emp_name, base_date, "PUT_WORK_RECORD (PAID_HOLIDAY)", resp, None)
                continue

            # --- 欠勤は PUT（is_absence） ---
            # 休日に「欠勤」を入れたくない運用なら、所定労働日のみに制限
            if is_absence:
                if absence_only_on_workday and holiday_based != "所定労働日":
                    _log(
                        staff_code,
                        emp_name,
                        base_date,
                        f"SKIP_ABSENCE (day_type={holiday_based})",
                        None,
                        "not a workday",
                    )
                    continue

                # 欠勤なのに実績がある場合は、上書き事故になりやすいので安全側にスキップ
                # （運用上「実績優先」がおすすめ。欠勤優先にしたいならこのifを削除）
                if has_actual:
                    _log(
                        staff_code,
                        emp_name,
                        base_date,
                        "SKIP_ABSENCE (has_actual)",
                        None,
                        "status=absence but has clock_in/out",
                    )
                    continue

                body = build_absence_work_record_body(
                    company_id,
                    base_date,
                    use_attendance_deduction=absence_use_attendance_deduction,
                )
                resp = client.put_work_record(employee_id, base_date, body)
                _log(staff_code, emp_name, base_date, "PUT_WORK_RECORD (ABSENCE)", resp, None)
                continue

            # ---- ここから通常日 ----
            # 実績がある日だけ同期（所定労働日も休日も同様）
            if not has_actual:
                continue

            # 実績からセグメント
            sdt, edt, bseg = normalize_day_segments(
                wd,
                rec["clock_in"],
                rec["clock_out"],
                int(rec.get("break_mins", 0) or 0),
            )

            # ===== POST time_clocks で打刻登録 =====
            events: List[Tuple[str, datetime]] = [("clock_in", sdt)]
            if bseg:
                b_start, b_end = bseg
                events += [("break_begin", b_start), ("break_end", b_end)]
            events += [("clock_out", edt)]

            for ttype, tdt in events:
                dt_str = tdt.strftime("%Y-%m-%d %H:%M:%S")
                resp = client.post_time_clock(employee_id, company_id, ttype, base_date, dt_str)
                _log(staff_code, emp_name, base_date, f"POST_TIME_CLOCK {ttype} (day_type={holiday_based})", resp, None)

    df = pd.DataFrame(results)
    csv_bytes = df_to_csv_bytes_utf8sig(df)
    return df, csv_bytes



# =========================
# Streamlit cache helpers
# =========================
@st.cache_resource
def make_client_cached() -> FreeeHrClient:
    client_id = os.environ.get("FREEE_CLIENT_ID", "").strip()
    client_secret = os.environ.get("FREEE_CLIENT_SECRET", "").strip()
    initial_refresh = os.environ.get("FREEE_REFRESH_TOKEN", "").strip()

    if not client_id or not client_secret:
        raise RuntimeError("FREEE_CLIENT_ID / FREEE_CLIENT_SECRET が未設定です（.env を確認）。")

    token = FreeeToken.load()
    if token is None:
        if not initial_refresh:
            raise RuntimeError("初回は環境変数 FREEE_REFRESH_TOKEN が必要です（token_store.json がまだ無い）。")
        token = FreeeToken(access_token="DUMMY", refresh_token=initial_refresh)
        FreeeHrClient(client_id, client_secret, token)._refresh()
    else:
        # .env に新しいリフレッシュトークンがあればそちらを優先（期限切れ token_store 対策）
        if initial_refresh and initial_refresh != token.refresh_token:
            token.refresh_token = initial_refresh
            token.save()

    return FreeeHrClient(client_id, client_secret, token)

@st.cache_data(show_spinner=False)
def read_jobcan_excels_cached(files: List[Tuple[str, bytes]]) -> Tuple[int, int, Dict[str, List[Dict[str, Any]]]]:
    return read_jobcan_excels(files)

@st.cache_data(show_spinner=False)
def fetch_employees_cached(company_id: int, year: Optional[int] = None, month: Optional[int] = None) -> List[Dict[str, Any]]:
    client = make_client_cached()
    all_items: List[Dict[str, Any]] = []
    offset = 0
    limit = 100
    while True:
        params: Dict[str, Any] = {"limit": limit, "offset": offset}
        if year is not None:
            params["year"] = year
        if month is not None:
            params["month"] = month
        r = client.request("GET", f"/companies/{company_id}/employees", params=params)
        r.raise_for_status()
        payload = r.json()
        if isinstance(payload, list):
            items = payload
        elif isinstance(payload, dict):
            items = next(
                (payload[k] for k in ["employees", "items", "data"] if k in payload and isinstance(payload[k], list)),
                ([payload] if "id" in payload else [])
            )
        else:
            items = []
        all_items.extend(items)
        if len(items) < limit:
            break
        offset += limit
    return all_items


# =========================
# Streamlit UI
# =========================
load_dotenv_simple(DOTENV_PATH, override=False)

st.set_page_config(page_title="freee人事労務 × ジョブカン勤怠 連携", layout="wide")

# ----- OAuth コールバック: URL に code が付いていればトークン取得 -----
try:
    qp = st.query_params
except Exception:
    qp = {}
if qp.get("code"):
    client_id = os.environ.get("FREEE_CLIENT_ID", "").strip()
    client_secret = os.environ.get("FREEE_CLIENT_SECRET", "").strip()
    redirect_uri = get_redirect_uri()
    if os.path.exists(OAUTH_STATE_PATH):
        try:
            with open(OAUTH_STATE_PATH, "r", encoding="utf-8") as f:
                saved = json.load(f)
                redirect_uri = saved.get("redirect_uri") or redirect_uri
        except Exception:
            pass
    if client_id and client_secret:
        try:
            exchange_code_for_token(
                client_id, client_secret, qp["code"], redirect_uri
            )
            if os.path.exists(OAUTH_STATE_PATH):
                os.remove(OAUTH_STATE_PATH)
            make_client_cached.clear()
            if hasattr(st.query_params, "clear"):
                st.query_params.clear()
            else:
                try:
                    st.experimental_set_query_params()
                except Exception:
                    pass
            st.success("freee と連携しました。このページを再読み込みしてください。")
        except Exception as e:
            st.error(f"トークン取得に失敗しました: {e}")
    else:
        st.error("FREEE_CLIENT_ID / FREEE_CLIENT_SECRET を .env に設定してください。")
    st.stop()

st.title("freee人事労務 × ジョブカン勤怠 連携（CSV固定列・2ステップ実行）")

with st.sidebar:
    # freee 認証（リフレッシュトークンが無い場合）
    client_id = os.environ.get("FREEE_CLIENT_ID", "").strip()
    client_secret = os.environ.get("FREEE_CLIENT_SECRET", "").strip()
    has_creds = bool(client_id and client_secret)
    if has_creds:
        st.subheader("freee 認証")
        st.caption("リフレッシュトークンが無い／期限切れのときは、下の「認証を開始」で freee にログインして取得します。")
        if "oauth_url" not in st.session_state:
            if st.button("認証を開始（URLを表示）"):
                state = secrets.token_urlsafe(16)
                redirect_uri = get_redirect_uri()
                with open(OAUTH_STATE_PATH, "w", encoding="utf-8") as f:
                    json.dump({"state": state, "redirect_uri": redirect_uri}, f)
                st.session_state["oauth_url"] = build_authorize_url(
                    client_id, redirect_uri, state
                )
                st.rerun()
        if st.session_state.get("oauth_url"):
            st.markdown(
                f"[**このリンクを開いて freee でログイン**]({st.session_state['oauth_url']})"
            )
            st.caption("認証後、このアプリの画面に戻ります。")
            st.caption("freee 開発者向けAPIで「リダイレクトURI」に " + get_redirect_uri() + " を登録してください。")
        st.divider()

    st.subheader("設定")
    do_schedule = st.checkbox("① スケジュールCSVを生成", value=True)
    do_sync = st.checkbox("② API同期（打刻＋有休PUT）", value=True)

    st.divider()
    fast_mode = False  # 常に削除→POST方式

    st.divider()
    st.caption("CSV列は固定（テンプレ不要）です")
    st.caption("休日区分は '祝日, 公休' / '祝日, 法休' などでも休日扱いします")
    st.caption("シフトが 0:00/0:00 の場合は休憩を必ずブランクにします")
    st.caption("B列が「有休」→ work_records PUT で paid_holiday=1 を登録します")

st.subheader("入力")
uploaded_excels = st.file_uploader("ジョブカン summary Excel（複数可）", type=["xlsx"], accept_multiple_files=True)

# セッション状態
for k in [
    "prepared",
    "year",
    "month",
    "staff_rows",
    "employee_map",
    "schedule_bytes",
    "schedule_df_preview",
]:
    if k not in st.session_state:
        st.session_state[k] = None
if "prepared" not in st.session_state or st.session_state["prepared"] is None:
    st.session_state["prepared"] = False

st.divider()
st.subheader("STEP1: CSV生成（ここで一旦停止）")

btn_prepare_disabled = (not uploaded_excels) or (not do_schedule)
if st.button("① CSVを生成して停止", type="primary", disabled=btn_prepare_disabled):
    try:
        client = make_client_cached()

        files = [(f.name, f.getvalue()) for f in uploaded_excels]
        year, month, staff_rows = read_jobcan_excels_cached(files)
        st.success(f"[OK] Excel読込完了：year={year} month={month} 対象従業員（シート数ベース）={len(staff_rows)}")

        employees_list = fetch_employees_cached(COMPANY_ID, year=year, month=month)
        employee_map = build_employee_map(employees_list)
        st.success(f"[OK] freee従業員取得：{len(employees_list)}件 / employee_numberマップ={len(employee_map)}件")

        missing = [code for code in staff_rows.keys() if code not in employee_map]
        if missing:
            st.warning(f"freee側に存在しない従業員番号（スキップ）: {missing[:20]}{'...' if len(missing)>20 else ''}")

        with st.expander("🔍 診断情報：freee従業員番号マップ（確認用）", expanded=bool(missing)):
            excel_codes = sorted(staff_rows.keys())
            freee_codes = sorted(employee_map.keys())
            matched = [c for c in excel_codes if c in employee_map]
            st.write(f"**Excelのスタッフコード一覧** ({len(excel_codes)}件): {excel_codes}")
            st.write(f"**freeeの従業員番号一覧** ({len(freee_codes)}件): {freee_codes}")
            st.write(f"**マッチ済み** ({len(matched)}件): {matched}")
            if missing:
                st.error(f"**未マッチ（スキップされる）** ({len(missing)}件): {missing}")
                st.caption("→ freee人事労務の管理画面で、上記の従業員番号が登録されているか確認してください。")

        schedule_df, schedule_bytes = build_schedule_csv_fixed(employee_map, staff_rows)

        st.session_state["prepared"] = True
        st.session_state["year"] = year
        st.session_state["month"] = month
        st.session_state["staff_rows"] = staff_rows
        st.session_state["employee_map"] = employee_map
        st.session_state["schedule_bytes"] = schedule_bytes
        st.session_state["schedule_df_preview"] = schedule_df.head(200)

        st.success(f"[OK] スケジュールCSV生成：rows={len(schedule_df)}（ここで停止します）")

    except Exception as e:
        st.session_state["prepared"] = False
        st.exception(e)

if st.session_state.get("prepared") and st.session_state.get("schedule_bytes"):
    st.write(f"準備済み: {st.session_state['year']}年{st.session_state['month']}月")
    st.dataframe(st.session_state["schedule_df_preview"], use_container_width=True)
    st.download_button(
        label="freee_schedule_import.csv をダウンロード",
        data=st.session_state["schedule_bytes"],
        file_name="freee_schedule_import.csv",
        mime="text/csv",
    )

st.divider()
st.subheader("STEP2: API同期（ボタンを押したら開始）")

btn_sync_disabled = (not do_sync) or (not st.session_state.get("prepared"))
confirm = st.checkbox(
    "（確認）CSVのダウンロードが完了しました。API同期を実行します。",
    value=False,
    disabled=btn_sync_disabled,
)

if st.button("② API同期を開始", type="primary", disabled=(btn_sync_disabled or (not confirm))):
    try:
        client = make_client_cached()
        staff_rows = st.session_state["staff_rows"]
        employee_map = st.session_state["employee_map"]

        # ---- 進捗UI（ここが追加）----
        total = sum(len(v) for v in staff_rows.values()) or 1
        progress_bar = st.progress(0.0)
        status_box = st.empty()

        def progress_cb(done: int, total: int, staff_code: str, emp_name: str, base_date: str) -> None:
            ratio = min(done / (total or 1), 1.0)
            progress_bar.progress(ratio)
            status_box.text(f"処理中 {done}/{total} : {staff_code} {emp_name} {base_date}")


        with st.spinner("API同期中..."):
            results_df, results_bytes = sync_time_clocks_and_work_records(
                client,
                employee_map,
                staff_rows,
                COMPANY_ID,
                fast_put_only=fast_mode,
                progress_cb=progress_cb,  # ★追加（sync関数側も対応が必要）
            )

        progress_bar.progress(1.0)
        status_box.text("API同期 完了")

        # 法定休日前日の24:00超
        alert_df = find_cross_midnight_before_statutory_holiday(employee_map, staff_rows)
        if not alert_df.empty:
            st.warning(f"法定休日の前日に 24:00超（翌日繰り越し）の実績打刻がある行: {len(alert_df)}件")
            st.dataframe(alert_df, use_container_width=True)
            st.download_button(
                label="法定休日前日の24時超打刻_一覧.csv をダウンロード",
                data=df_to_csv_bytes_utf8sig(alert_df),
                file_name="statutory_holiday_prevday_cross_midnight.csv",
                mime="text/csv",
            )
        else:
            st.info("法定休日の前日に 24:00超（翌日繰り越し）の打刻は見つかりませんでした。")

        st.success(f"[OK] 同期ログ生成：rows={len(results_df)} / mode={'FAST(PUTのみ)' if fast_mode else 'STRICT(確認してPOST)'}")
        st.dataframe(results_df.head(300), use_container_width=True)
        st.download_button(
            label="sync_results.csv をダウンロード",
            data=results_bytes,
            file_name="sync_results.csv",
            mime="text/csv",
        )

    except Exception as e:
        st.exception(e)
