# app.py
# -*- coding: utf-8 -*-
"""
勤怠CSV（パート12/1～）から勤怠項目を分解し、
従業員情報CSV（employee_exports）から時給（早番/遅番）を自動生成して、
従業員別に給与計算まで行う（Streamlit版）

- file_uploaderで2つのCSVを受け取り
- 画面で「時給マスタ」「日別明細」「従業員別集計_給与」を確認
- Excel（3シート）をダウンロード
"""

from __future__ import annotations

import math
import re
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl.styles import PatternFill

# =========================================================
# 設定
# =========================================================
TARGET_WAGE_CATEGORY = "パート12/1～"

# 早番／遅番の境界：16:30
CUTOFF_MIN = 16 * 60 + 30

# 割増（固定）
RATE_OT = 1.25   # 時間外
RATE_HOL = 1.35  # 休日
RATE_NIGHT_PREMIUM = 0.25  # 深夜手当（手当分のみ）
# ※法定内残業は 1.0

# =========================================================
# 勤怠CSVの列名（あなたのCSVヘッダに合わせる）
# =========================================================
K_COL_WAGE = "勤務・賃金"
K_COL_DATE = "日付"
K_COL_EMP_NO = "従業員番号"
K_COL_NAME = "氏名"
K_COL_DEPT = "部門"

K_COL_DAYTYPE = "勤務日種別"          # 所定休日
K_COL_ATT_TYPE = "勤怠種別"           # 有給休暇

K_COL_START = "出勤時刻"
K_COL_END = "退勤時刻"
K_COL_PLAN_START = "勤務予定開始時刻"
K_COL_PLAN_END = "勤務予定退勤時刻"

K_COL_BREAK = "休憩時間"
K_COL_SCHEDULED_WORK = "所定内労働時間"  # 有休の長さはここから

K_COL_LEGALIN_OT = "法定内残業時間"    # O列（法定内）
K_COL_OUTSIDE_OT = "時間外労働時間"    # P列（時間外）
K_COL_NIGHT_TOTAL = "深夜労働時間"     # S列（深夜）

# =========================================================
# 従業員exports CSVの列名（employee_exports）
# =========================================================
E_COL_EMP_NO = "従業員番号"
E_COL_RETIRE = "退職日"
E_COL_EMPLOY = "雇用形態"
E_COL_BASE_WAGE = "基本給"
E_TW_NAME_TMPL = "時間帯ごとの時給{n} 名前"
E_TW_AMT_TMPL = "時間帯ごとの時給{n} 金額"


# =========================================================
# I/O（Streamlitアップロード対応）
# =========================================================
def read_csv_robust_from_bytes(b: bytes) -> pd.DataFrame:
    encodings = ["utf-8-sig", "utf-8", "cp932", "shift_jis"]
    last_err = None
    for enc in encodings:
        try:
            return pd.read_csv(BytesIO(b), encoding=enc)
        except Exception as e:
            last_err = e
    raise RuntimeError(f"CSVの読み込みに失敗しました / last_error={last_err}")


def safe_filename(s: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "_", str(s)).strip()


# =========================================================
# パース
# =========================================================
def parse_hhmm_to_min(v: Any) -> Optional[int]:
    if pd.isna(v):
        return None
    s = str(v).strip()
    if s == "" or s.lower() in ("nat", "nan", "none"):
        return None
    m = re.match(r"^(\d{1,2}):(\d{2})(?::(\d{2}))?$", s)
    if not m:
        return None
    h = int(m.group(1))
    mi = int(m.group(2))
    return h * 60 + mi


def parse_duration_minutes(v: Any) -> int:
    mm = parse_hhmm_to_min(v)
    return int(mm or 0)


def parse_break_minutes(v: Any) -> int:
    if pd.isna(v):
        return 0
    s = str(v).strip()
    if s == "":
        return 0

    mm = parse_hhmm_to_min(s)
    if mm is not None:
        return mm

    try:
        f = float(s)
        if f < 10:
            return int(round(f * 60))
        return int(round(f))
    except Exception:
        return 0


def minutes_to_hhmm(mins: int) -> str:
    mins = int(round(mins))
    sign = "-" if mins < 0 else ""
    mins = abs(mins)
    return f"{sign}{mins // 60:02d}:{mins % 60:02d}"


def to_float(v: Any) -> float:
    if pd.isna(v):
        return float("nan")
    s = str(v).strip()
    if s == "" or s.lower() == "nan":
        return float("nan")
    s = s.replace(",", "")
    try:
        return float(s)
    except Exception:
        return float("nan")


def normalize_empno_to_int(s: pd.Series) -> pd.Series:
    """
    従業員番号の正規化（重要）
    - 300344.0 → 300344
    - "000300344" → 300344
    - "社員300344" → 300344
    """
    num = pd.to_numeric(s, errors="coerce")
    out = pd.Series(pd.NA, index=s.index, dtype="Int64")

    m = num.notna()
    if m.any():
        out.loc[m] = num.loc[m].round(0).astype("Int64")

    m2 = ~m
    if m2.any():
        txt = s.loc[m2].astype(str).str.strip()
        txt = txt.str.replace(r"\.0+$", "", regex=True)
        txt = txt.str.replace(r"\D", "", regex=True)
        out.loc[m2] = pd.to_numeric(txt, errors="coerce").astype("Int64")

    return out.fillna(0).astype(int)


# =========================================================
# セグメント（早番/遅番判定）
# =========================================================
def build_segments(start_min: Optional[int], end_min: Optional[int]) -> List[List[Any]]:
    """
    [start,end) を、16:30境界で分割したセグメント配列にする
    seg = [a, b, is_early]
    """
    if start_min is None or end_min is None:
        return []

    s = start_min
    e = end_min
    if e < s:
        e += 1440

    boundaries = {s, e}
    day0 = s // 1440
    day1 = e // 1440
    for d in range(int(day0), int(day1) + 1):
        boundaries.add(d * 1440 + CUTOFF_MIN)

    cuts = sorted(b for b in boundaries if s < b < e)
    points = [s] + cuts + [e]

    segs: List[List[Any]] = []
    for a, b in zip(points[:-1], points[1:]):
        mid = (a + b) / 2
        local = int(mid) % 1440
        is_early = local < CUTOFF_MIN
        segs.append([a, b, is_early])
    return segs


def apply_break_rule(segs: List[List[Any]], break_min: int) -> List[List[Any]]:
    """
    休憩控除：
      - 基本は早番側から控除
      - ただし早番で控除しきれない場合は「早番からは控除しない」で休憩全額を遅番から控除
    """
    if break_min <= 0 or not segs:
        return segs

    early_total = sum((b - a) for a, b, is_early in segs if is_early)

    def deduct_from(segs_in: List[List[Any]], target_is_early: bool, minutes: int) -> List[List[Any]]:
        remaining = minutes
        new: List[List[Any]] = []
        for a, b, is_early in segs_in:
            if is_early == target_is_early and remaining > 0:
                dur = b - a
                if dur <= remaining:
                    remaining -= dur
                    continue
                a2 = a + remaining
                remaining = 0
                new.append([a2, b, is_early])
            else:
                new.append([a, b, is_early])
        return new

    if early_total >= break_min:
        return deduct_from(segs, True, break_min)

    return deduct_from(segs, False, break_min)


def trim_from_start(segs: List[List[Any]], minutes: int) -> List[List[Any]]:
    """先頭から minutes 分だけ残す（有休: N分で確定）"""
    if minutes <= 0 or not segs:
        return []
    remaining = minutes
    out: List[List[Any]] = []
    for a, b, is_early in segs:
        if remaining <= 0:
            break
        dur = b - a
        if dur <= remaining:
            out.append([a, b, is_early])
            remaining -= dur
        else:
            out.append([a, a + remaining, is_early])
            remaining = 0
    return out


def take_from_end(segs: List[List[Any]], minutes: int) -> Tuple[List[List[Any]], List[List[Any]]]:
    """末尾から minutes 分だけ切り出す（残り, 切り出し）"""
    if minutes <= 0 or not segs:
        return segs, []
    remaining = minutes
    extracted_rev: List[List[Any]] = []
    rem_rev: List[List[Any]] = []

    for a, b, is_early in reversed(segs):
        dur = b - a
        if remaining <= 0:
            rem_rev.append([a, b, is_early])
            continue
        if dur <= remaining:
            extracted_rev.append([a, b, is_early])
            remaining -= dur
        else:
            split = b - remaining
            extracted_rev.append([split, b, is_early])
            rem_rev.append([a, split, is_early])
            remaining = 0

    return list(reversed(rem_rev)), list(reversed(extracted_rev))


def sum_minutes_by_early(segs: List[List[Any]]) -> Tuple[int, int]:
    early = 0
    late = 0
    for a, b, is_early in segs:
        dur = b - a
        if is_early:
            early += dur
        else:
            late += dur
    return early, late


# =========================================================
# 従業員exports → 時給マスタ（自動生成）
# =========================================================
def build_wage_master(emp_df: pd.DataFrame) -> pd.DataFrame:
    """
    雇用形態や退職日で除外しない（勤怠に出てくる＝給与計算対象になりうるため）
    時給設定（基本給 or 時給1〜5のいずれか）がある行だけを対象にする。
    """
    tmp = emp_df.copy()

    if E_COL_EMP_NO not in tmp.columns or E_COL_BASE_WAGE not in tmp.columns:
        raise RuntimeError("従業員CSVに必要な列（従業員番号/基本給）がありません。")

    tmp[E_COL_EMP_NO] = tmp[E_COL_EMP_NO].astype(str).str.strip()
    tmp["_empno_int"] = normalize_empno_to_int(tmp[E_COL_EMP_NO])

    base = tmp[E_COL_BASE_WAGE].map(to_float)

    early = base.copy()
    late = base.copy()

    any_timeband_amt = pd.Series([False] * len(tmp), index=tmp.index)

    for i in range(1, 6):
        name_col = E_TW_NAME_TMPL.format(n=i)
        amt_col = E_TW_AMT_TMPL.format(n=i)
        if name_col not in tmp.columns or amt_col not in tmp.columns:
            continue

        nm = tmp[name_col].astype(str).fillna("").str.strip()
        am = tmp[amt_col].map(to_float)

        any_timeband_amt = any_timeband_amt | am.notna()

        # 早番/遅番をスロット問わず反映
        early = early.where(nm != "早番", am)
        late = late.where(nm != "遅番", am)

        # 逆転ロジック（あなたの仕様）
        late = late.where(nm != "早番", base)   # 時間帯=早番 → 基本給は遅番
        early = early.where(nm != "遅番", base) # 時間帯=遅番 → 基本給は早番

    # 時給設定がある人だけ残す
    has_wage = base.notna() | any_timeband_amt
    tmp = tmp[has_wage].copy()
    early = early.loc[tmp.index]
    late = late.loc[tmp.index]

    # 片側だけ取れたときのフォールバック
    early = early.fillna(late)
    late = late.fillna(early)

    out = pd.DataFrame({
        "従業員番号": tmp[E_COL_EMP_NO],
        "_empno_int": tmp["_empno_int"],
        "早番時給": early,
        "遅番時給": late,
    })

    # 従業員番号が同じ行が複数ある場合、最後の行を採用
    out = out.drop_duplicates(subset=["_empno_int"], keep="last")

    return out


# =========================================================
# 勤怠1行 → 分解（min）
# =========================================================
def compute_one_kintai_row(row: pd.Series) -> Dict[str, Any]:
    day_type = str(row.get(K_COL_DAYTYPE, "")).strip()
    holiday = (day_type == "所定休日")

    att_type = str(row.get(K_COL_ATT_TYPE, "")).strip()
    # CSVの表記ゆれ対策：「有休」「有給休暇」など
    att_type_norm = re.sub(r"\s+", "", att_type)
    is_paid = ("有休" in att_type_norm) or ("有給" in att_type_norm) or ("年休" in att_type_norm)

    legalin_ot_min = parse_duration_minutes(row.get(K_COL_LEGALIN_OT, "00:00"))
    outside_ot_min = parse_duration_minutes(row.get(K_COL_OUTSIDE_OT, "00:00"))
    night_total_min = parse_duration_minutes(row.get(K_COL_NIGHT_TOTAL, "00:00"))

    # 勤務時間帯（配賦用）
    if is_paid:
        s = parse_hhmm_to_min(row.get(K_COL_PLAN_START))
        e = parse_hhmm_to_min(row.get(K_COL_PLAN_END))
        source = "有休(予定+所定内)"

        # 予定でセグメント化
        segs = build_segments(s, e)

        # --- 有休の実働分（休憩控除後）を、予定開始/退勤から確定させる ---
        # CSVの「所定内労働時間」が休憩控除前（=予定総時間と同値）になっているケースがあるため、
        # 予定総時間と休憩から「実働分」を作り、それを早番/遅番へ配賦する。
        base_len_col = parse_duration_minutes(row.get(K_COL_SCHEDULED_WORK, "00:00"))

        # --- 有休の休憩分を推定して、休憩控除ルールを適用する ---
        # 1) まず休憩時間列が入っていればそれを使う
        br = parse_break_minutes(row.get(K_COL_BREAK))

        # 2) 空/0なら、(予定の長さ - 所定内労働時間) で休憩を推定
        plan_total = 0
        if br <= 0 and s is not None and e is not None:
            plan_total = e - s
            if plan_total < 0:
                plan_total += 1440
            br = max(plan_total - base_len_col, 0)
        elif s is not None and e is not None:
            plan_total = e - s
            if plan_total < 0:
                plan_total += 1440

        # 3) 有休の実働分（休憩控除後）を確定
        # - 所定内労働時間が0/空なら、予定総時間-休憩
        # - 所定内労働時間が予定総時間と同じ（休憩控除前）なら、予定総時間-休憩
        # - それ以外は所定内労働時間を優先
        if base_len_col <= 0 and plan_total > 0:
            base_len = max(plan_total - br, 0)
        elif plan_total > 0 and base_len_col == plan_total and br > 0:
            base_len = max(plan_total - br, 0)
        else:
            base_len = base_len_col

        # 休憩控除（原則 早番から、足りなければ全額 遅番）
        segs = apply_break_rule(segs, br)

        # 念のため、所定内労働時間(N)に合わせて切る（推定休憩がズレても安定）
        segs = trim_from_start(segs, base_len)

        # 有休は深夜を取らない運用（深夜はCSVのS列集計なので通常0想定）
    else:
        s = parse_hhmm_to_min(row.get(K_COL_START))
        e = parse_hhmm_to_min(row.get(K_COL_END))
        source = "実績"
        segs = build_segments(s, e)
        br = parse_break_minutes(row.get(K_COL_BREAK))
        segs = apply_break_rule(segs, br)


    total_min = sum((b - a) for a, b, _ in segs)
    early_total, late_total = sum_minutes_by_early(segs)

    # 休日：休日の「昼（深夜以外）」も早番/遅番で支給
    if holiday:
        hol_night = night_total_min  # 深夜は勤怠S列を採用
        hol_day = max(total_min - hol_night, 0)

        # 深夜は必ず遅番なので、昼の配賦は「早番→遅番」の順で充当
        hol_day_early = min(early_total, hol_day)
        hol_day_late = max(hol_day - hol_day_early, 0)

        return {
            "勤務区分": source,

            # 平日枠は0
            "早番時間(基本)_min": 0,
            "遅番時間(基本)_min": 0,
            "法定内残業_早番_min": 0,
            "法定内残業_遅番_min": 0,
            "時間外労働_早番_min": 0,
            "時間外労働_遅番_min": 0,

            # 深夜（全日）
            "深夜労働_min": hol_night,

            # 互換列（残しておく）
            "休日労働_min": hol_day,
            "休日深夜_min": hol_night,

            # 休日の早番/遅番（給与計算はこれを使う）
            "休日労働_早番_min": hol_day_early,
            "休日労働_遅番_min": hol_day_late,
            "休日深夜_遅番_min": hol_night,
        }

    # 平日：末尾から「時間外→法定内」を切り出し、残りが基本
    segs_rem = segs
    segs_rem, segs_outside = take_from_end(segs_rem, outside_ot_min)  # 時間外
    segs_rem, segs_legalin = take_from_end(segs_rem, legalin_ot_min)  # 法定内
    segs_basic = segs_rem

    early_b, late_b = sum_minutes_by_early(segs_basic)
    early_li, late_li = sum_minutes_by_early(segs_legalin)
    early_ot, late_ot = sum_minutes_by_early(segs_outside)

    return {
        "勤務区分": source,
        "早番時間(基本)_min": early_b,
        "遅番時間(基本)_min": late_b,
        "法定内残業_早番_min": early_li,
        "法定内残業_遅番_min": late_li,
        "時間外労働_早番_min": early_ot,
        "時間外労働_遅番_min": late_ot,

        # 深夜（S列集計）
        "深夜労働_min": night_total_min,

        "休日労働_min": 0,
        "休日深夜_min": 0,
        "休日労働_早番_min": 0,
        "休日労働_遅番_min": 0,
        "休日深夜_遅番_min": 0,
    }


# =========================================================
# 給与計算（端数：賃金（円）の最後にだけceil）
# =========================================================
def calc_amount_from_minutes(rate_yen_per_hour: float, minutes: int) -> int:
    if minutes <= 0 or pd.isna(rate_yen_per_hour):
        return 0
    return int(math.ceil(float(rate_yen_per_hour) * float(minutes) / 60.0))


def calc_summary_pay(summary: pd.DataFrame, wage_master: pd.DataFrame) -> pd.DataFrame:
    summary = summary.copy()
    summary["_empno_int"] = normalize_empno_to_int(summary["従業員番号"])

    wm = wage_master.copy()
    wm["_empno_int"] = normalize_empno_to_int(wm["従業員番号"])
    wm = wm.drop_duplicates("_empno_int")[["_empno_int", "早番時給", "遅番時給"]]

    summary = summary.merge(wm, how="left", on="_empno_int")

    summary["割増率_時間外"] = RATE_OT
    summary["割増率_休日"] = RATE_HOL

    summary["時給未取得"] = summary["早番時給"].isna() | summary["遅番時給"].isna()

    def m(col: str) -> pd.Series:
        return summary[col].fillna(0).astype(int)

    m_eb   = m("早番時間(基本)_min")
    m_lb   = m("遅番時間(基本)_min")
    m_li_e = m("法定内残業_早番_min")
    m_li_l = m("法定内残業_遅番_min")
    m_ot_e = m("時間外労働_早番_min")
    m_ot_l = m("時間外労働_遅番_min")
    m_n    = m("深夜労働_min")

    m_h_e  = m("休日労働_早番_min")
    m_h_l  = m("休日労働_遅番_min")
    m_hn_l = m("休日深夜_遅番_min")

    early_w = pd.to_numeric(summary["早番時給"], errors="coerce").fillna(0.0)
    late_w  = pd.to_numeric(summary["遅番時給"], errors="coerce").fillna(0.0)
    summary["基準時給_推定"] = ((early_w + late_w) / 2.0).round(2)

    # 単価（float）※ここでは端数処理しない
    early_basic_rate = early_w
    late_basic_rate  = late_w

    early_legalin_rate = early_w  # 1.0倍
    late_legalin_rate  = late_w   # 1.0倍

    early_ot_rate = early_w * RATE_OT
    late_ot_rate  = late_w * RATE_OT

    early_hol_rate = early_w * RATE_HOL
    late_hol_rate  = late_w * RATE_HOL

    # 深夜手当は必ず遅番時給×0.25
    night_premium_rate = late_w * RATE_NIGHT_PREMIUM

    # 金額（円）※金額だけceil
    summary["金額_早番基本"] = [calc_amount_from_minutes(r, mins) for r, mins in zip(early_basic_rate, m_eb)]
    summary["金額_遅番基本"] = [calc_amount_from_minutes(r, mins) for r, mins in zip(late_basic_rate,  m_lb)]

    summary["金額_法定内残業_早番"] = [calc_amount_from_minutes(r, mins) for r, mins in zip(early_legalin_rate, m_li_e)]
    summary["金額_法定内残業_遅番"] = [calc_amount_from_minutes(r, mins) for r, mins in zip(late_legalin_rate,  m_li_l)]

    summary["金額_時間外_早番"] = [calc_amount_from_minutes(r, mins) for r, mins in zip(early_ot_rate, m_ot_e)]
    summary["金額_時間外_遅番"] = [calc_amount_from_minutes(r, mins) for r, mins in zip(late_ot_rate,  m_ot_l)]

    # 休日（早番/遅番）
    summary["金額_休日労働_早番"] = [calc_amount_from_minutes(r, mins) for r, mins in zip(early_hol_rate, m_h_e)]
    summary["金額_休日労働_遅番"] = [calc_amount_from_minutes(r, mins) for r, mins in zip(late_hol_rate,  m_h_l)]

    # 休日深夜：休日割増（遅番時給×1.35）＋深夜手当（遅番時給×0.25）
    summary["金額_休日深夜"] = [calc_amount_from_minutes(r, mins) for r, mins in zip(late_hol_rate, m_hn_l)]
    summary["金額_休日深夜_深夜手当"] = [calc_amount_from_minutes(r, mins) for r, mins in zip(night_premium_rate, m_hn_l)]

    # 深夜手当（平日深夜も含む：S列の深夜分に対して遅番時給×0.25）
    summary["金額_深夜手当"] = [calc_amount_from_minutes(r, mins) for r, mins in zip(night_premium_rate, m_n)]

    amt_cols = [
        "金額_早番基本", "金額_遅番基本",
        "金額_法定内残業_早番", "金額_法定内残業_遅番",
        "金額_時間外_早番", "金額_時間外_遅番",
        "金額_休日労働_早番", "金額_休日労働_遅番",
        "金額_休日深夜", "金額_休日深夜_深夜手当",
        "金額_深夜手当",
    ]
    summary["金額_合計"] = summary[amt_cols].fillna(0).sum(axis=1).astype(int)

    summary = summary.drop(columns=["_empno_int"])
    return summary


# =========================================================
# Excel出力（BytesIO）
# =========================================================
_YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def _apply_highlight(ws, row_positions: List[int]) -> None:
    n_cols = ws.max_column
    for pos in row_positions:
        excel_row = pos + 2  # +1 for header, +1 for 1-based
        for col in range(1, n_cols + 1):
            ws.cell(row=excel_row, column=col).fill = _YELLOW_FILL


def build_excel_bytes(
    detail: pd.DataFrame,
    summary_pay: pd.DataFrame,
    detail_highlight_rows: List[int] | None = None,
    summary_highlight_rows: List[int] | None = None,
) -> bytes:
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        # 従業員別集計_給与を先頭シートに
        summary_pay.to_excel(writer, index=False, sheet_name="従業員別集計_給与")
        detail.to_excel(writer, index=False, sheet_name="日別明細")

        if summary_highlight_rows:
            _apply_highlight(writer.sheets["従業員別集計_給与"], summary_highlight_rows)
        if detail_highlight_rows:
            _apply_highlight(writer.sheets["日別明細"], detail_highlight_rows)

    return bio.getvalue()


# =========================================================
# Streamlit UI
# =========================================================
def run_app() -> None:
    st.set_page_config(page_title="勤怠CSV→給与計算（早番/遅番）", layout="wide")
    st.title("勤怠CSV → 給与計算（早番/遅番・休日・深夜）")

    with st.expander("仕様（概要）", expanded=False):
        st.markdown(
            f"""
- 対象：勤怠CSVの「勤務・賃金」= **{TARGET_WAGE_CATEGORY}**
- 早番/遅番境界：**16:30**
- 休憩控除：
  - 基本は早番側から控除
  - 早番で控除しきれない場合は「早番から控除しない」で **休憩全額を遅番から控除**
- 休日：勤務日種別=所定休日
  - 休日労働も早番/遅番に分解して **×{RATE_HOL}**
  - 休日深夜：休日割増（遅番×{RATE_HOL}）＋深夜手当（遅番×{RATE_NIGHT_PREMIUM}）
- 深夜手当：必ず **遅番時給×{RATE_NIGHT_PREMIUM}**
- 端数処理：最後の「金額（円）」だけ **ceil**
"""
        )

    col1, col2 = st.columns(2)
    with col1:
        k_file = st.file_uploader("① 勤怠CSVをアップロード", type=["csv"], accept_multiple_files=False)
    with col2:
        e_file = st.file_uploader("② 従業員情報CSV（employee_exports）をアップロード", type=["csv"], accept_multiple_files=False)

    if not k_file or not e_file:
        st.info("上の2つのCSVをアップロードしてください。")
        return

    # 読み込み
    try:
        k = read_csv_robust_from_bytes(k_file.getvalue())
        e = read_csv_robust_from_bytes(e_file.getvalue())
    except Exception as ex:
        st.error(f"CSV読み込みに失敗しました: {ex}")
        return

    # 必須列チェック（勤怠）
    k_required = [
        K_COL_WAGE, K_COL_DATE, K_COL_EMP_NO, K_COL_NAME, K_COL_DEPT,
        K_COL_DAYTYPE, K_COL_ATT_TYPE,
        K_COL_START, K_COL_END, K_COL_PLAN_START, K_COL_PLAN_END,
        K_COL_BREAK, K_COL_SCHEDULED_WORK,
        K_COL_LEGALIN_OT, K_COL_OUTSIDE_OT, K_COL_NIGHT_TOTAL
    ]
    missing_k = [c for c in k_required if c not in k.columns]
    if missing_k:
        st.error(f"勤怠CSVに必要な列がありません: {missing_k}")
        return

    # 必須列チェック（従業員）
    missing_e = [c for c in [E_COL_EMP_NO, E_COL_BASE_WAGE] if c not in e.columns]
    if missing_e:
        st.error(f"従業員CSVに必要な列がありません: {missing_e}")
        return

    # 実行ボタン（重い処理の誤爆防止）
    if not st.button("計算を実行", type="primary"):
        st.stop()

    # 1) 時給マスタ自動生成
    try:
        wage_master = build_wage_master(e)
    except Exception as ex:
        st.error(f"時給マスタ作成に失敗しました: {ex}")
        return

    # 2) 勤怠：対象のみ抽出
    k = k[k[K_COL_WAGE].astype(str).str.strip() == TARGET_WAGE_CATEGORY].copy()
    if k.empty:
        st.warning(f"勤怠CSVに対象（{TARGET_WAGE_CATEGORY}）がありません。")
        return

    # 3) 日別明細：分解
    k[K_COL_EMP_NO] = k[K_COL_EMP_NO].astype(str).str.strip()
    calc = k.apply(compute_one_kintai_row, axis=1, result_type="expand")
    out = pd.concat([k, calc], axis=1)

    # 4) 日別明細に時給を結合（表示用）
    out["_empno_int"] = normalize_empno_to_int(out[K_COL_EMP_NO])
    wm2 = wage_master.copy()
    out = out.merge(wm2[["_empno_int", "早番時給", "遅番時給"]], how="left", on="_empno_int")

    # 5) 従業員別集計（min合計）
    sum_min_cols = [
        "早番時間(基本)_min", "遅番時間(基本)_min",
        "法定内残業_早番_min", "法定内残業_遅番_min",
        "時間外労働_早番_min", "時間外労働_遅番_min",
        "深夜労働_min",
        "休日労働_min", "休日深夜_min",
        "休日労働_早番_min", "休日労働_遅番_min", "休日深夜_遅番_min",
    ]
    summary = out.groupby([K_COL_EMP_NO, K_COL_NAME, K_COL_DEPT], as_index=False)[sum_min_cols].sum()

    # 6) 給与計算
    summary_pay = calc_summary_pay(summary, wage_master)

    # 7) 表示用（時間 hh:mm）
    time_cols = [
        ("早番時間(基本)", "早番時間(基本)_min"),
        ("遅番時間(基本)", "遅番時間(基本)_min"),
        ("法定内残業_早番", "法定内残業_早番_min"),
        ("法定内残業_遅番", "法定内残業_遅番_min"),
        ("時間外労働_早番", "時間外労働_早番_min"),
        ("時間外労働_遅番", "時間外労働_遅番_min"),
        ("深夜労働", "深夜労働_min"),
        ("休日労働_早番", "休日労働_早番_min"),
        ("休日労働_遅番", "休日労働_遅番_min"),
        ("休日深夜", "休日深夜_遅番_min"),
    ]
    for disp, raw in time_cols:
        if raw in out.columns:
            out[disp] = out[raw].fillna(0).astype(int).map(minutes_to_hhmm)
        if raw in summary_pay.columns:
            summary_pay[disp] = summary_pay[raw].fillna(0).astype(int).map(minutes_to_hhmm)

    # 8) 従業員別集計_給与：min列を落として hh:mm のみ残す、F〜U列（割増率・金額列）も削除
    min_cols_summary = [c for c in summary_pay.columns if c.endswith("_min")]
    summary_pay_disp = summary_pay.drop(columns=min_cols_summary)
    drop_cols_summary = [
        "割増率_時間外", "割増率_休日", "時給未取得", "基準時給_推定",
        "金額_早番基本", "金額_遅番基本",
        "金額_法定内残業_早番", "金額_法定内残業_遅番",
        "金額_時間外_早番", "金額_時間外_遅番",
        "金額_休日労働_早番", "金額_休日労働_遅番",
        "金額_休日深夜", "金額_休日深夜_深夜手当",
        "金額_深夜手当", "金額_合計",
    ]
    summary_pay_disp = summary_pay_disp.drop(columns=[c for c in drop_cols_summary if c in summary_pay_disp.columns])

    # 9) ダウンロード用Excel構築
    detail_cols = [
        K_COL_EMP_NO, K_COL_NAME, K_COL_DEPT, K_COL_DATE, K_COL_DAYTYPE, K_COL_ATT_TYPE, "勤務区分",
        K_COL_START, K_COL_END, K_COL_PLAN_START, K_COL_PLAN_END, K_COL_BREAK, K_COL_SCHEDULED_WORK,
        K_COL_LEGALIN_OT, K_COL_OUTSIDE_OT, K_COL_NIGHT_TOTAL,
        "早番時給", "遅番時給",
        "早番時間(基本)", "遅番時間(基本)",
        "法定内残業_早番", "法定内残業_遅番",
        "時間外労働_早番", "時間外労働_遅番",
        "深夜労働",
        "休日労働_早番", "休日労働_遅番", "休日深夜",
    ]
    detail = out[[c for c in detail_cols if c in out.columns]].copy()
    detail[K_COL_EMP_NO] = normalize_empno_to_int(detail[K_COL_EMP_NO])
    summary_pay_disp[K_COL_EMP_NO] = normalize_empno_to_int(summary_pay_disp[K_COL_EMP_NO])

    # 日別明細を従業員別集計_給与と同じ従業員順に並び替え（従業員内は日付順）
    emp_order = {empno: i for i, empno in enumerate(summary_pay_disp[K_COL_EMP_NO])}
    detail["_sort_key"] = detail[K_COL_EMP_NO].map(emp_order)
    detail = detail.sort_values(["_sort_key", K_COL_DATE]).drop(columns=["_sort_key"]).reset_index(drop=True)

    # 黄色ハイライト条件（有休行のみ対象）
    #   1. 16:30跨ぎ：早番・遅番の両方に時間がある
    #   2. 遅番の人（遅番時給 > 早番時給）が早番時間帯に有休
    _is_paid_row = detail["勤務区分"] == "有休(予定+所定内)"
    _has_early   = detail["早番時間(基本)"].fillna("00:00") != "00:00"
    _has_late    = detail["遅番時間(基本)"].fillna("00:00") != "00:00"
    _late_w      = pd.to_numeric(detail["遅番時給"], errors="coerce")
    _early_w     = pd.to_numeric(detail["早番時給"], errors="coerce")
    _is_late_worker = _late_w > _early_w

    paid_mask = _is_paid_row & (
        (_has_early & _has_late)          # 16:30跨ぎ
        | (_has_early & _is_late_worker)  # 遅番の人が早番時間帯に有休
    )
    detail_highlight_rows = [int(i) for i in paid_mask.values.nonzero()[0]]

    # 同対象者の従業員番号で従業員別集計_給与の行もハイライト
    highlight_empnos = set(detail.loc[paid_mask, K_COL_EMP_NO].tolist())
    summary_highlight_rows = [
        i for i, v in enumerate(summary_pay_disp[K_COL_EMP_NO])
        if v in highlight_empnos
    ]

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_cat = safe_filename(TARGET_WAGE_CATEGORY)
    out_name = f"{safe_filename(k_file.name.rsplit('.', 1)[0])}_給与計算_{safe_cat}_{stamp}.xlsx"

    try:
        excel_bytes = build_excel_bytes(detail, summary_pay_disp, detail_highlight_rows, summary_highlight_rows)
    except Exception as ex:
        st.error(f"Excel生成に失敗しました: {ex}")
        return

    # ---- 画面表示 ----
    st.success("計算が完了しました。下で結果を確認し、Excelをダウンロードできます。")

    nf_w = int(summary_pay["時給未取得"].sum()) if "時給未取得" in summary_pay.columns else 0
    if nf_w:
        st.warning(f"時給未取得が {nf_w} 人あります（従業員番号の一致を確認）")
        show_cols = [c for c in ["従業員番号", "氏名", "部門"] if c in summary_pay.columns]
        st.dataframe(summary_pay.loc[summary_pay["時給未取得"], show_cols].drop_duplicates(), use_container_width=True)

    tabs = st.tabs(["日別明細", "従業員別集計_給与"])
    with tabs[0]:
        st.dataframe(detail, use_container_width=True)
    with tabs[1]:
        st.dataframe(summary_pay_disp, use_container_width=True)

    st.download_button(
        label="Excelをダウンロード（2シート）",
        data=excel_bytes,
        file_name=out_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )


if __name__ == "__main__":
    run_app()
