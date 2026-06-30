import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime, time, date, timedelta
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import jpholiday
import re

# ─────────────────────────────────────────────
# 定数
# ─────────────────────────────────────────────
MONTHLY_STD_HOURS = 171.28  # 21.41日 × 8h


def ceil_int(x) -> int:
    """金額丸め: 切り上げ（手当・残業など）。微小な浮動小数点誤差をクリーンアップしてから切り上げる。"""
    if x is None:
        return 0
    cleaned = round(x, 4)  # 4桁丸めで浮動小数点誤差を除去
    return math.ceil(cleaned)


def floor_int(x) -> int:
    """金額丸め: 切り捨て（控除類）。"""
    if x is None:
        return 0
    cleaned = round(x, 4)
    return math.floor(cleaned)


def amount_by_min(rate: float, minutes: int, factor: float = 1.0, mode: str = "ceil") -> int:
    """rate(円/h) × minutes(分) / 60 × factor を計算し、丸める。

    分単位整数を使うことで浮動小数点誤差を最小化。"""
    amount = rate * minutes / 60 * factor
    if mode == "floor":
        return floor_int(amount)
    return ceil_int(amount)

STORE_FORMAT = {
    "コメダ珈琲店　旭川買物公園通り店": "asahikawa",
    "コメダ珈琲店　千歳北信濃店":     "chitose",
    "コメダ珈琲店　すすきの店":       "standard",
    "焼肉ホルモン　おはこ忠和店":     "standard",
}

ATTENDANCE_TYPES_PRESENT = {"通常勤務", "有休", "有給特休", "振替出勤"}
ABSENCE_TYPES = {"欠勤"}
ACTUAL_WORK_TYPES = {"通常勤務"}

# ─────────────────────────────────────────────
# 時間変換
# ─────────────────────────────────────────────
def to_minutes(val) -> int:
    """時間値を整数の総分数に変換（浮動小数点誤差を回避）"""
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return 0
    if isinstance(val, timedelta):
        return round(val.total_seconds() / 60)
    if isinstance(val, time):
        return val.hour * 60 + val.minute + round(val.second / 60)
    if isinstance(val, str):
        val = val.strip()
        if not val or val == "00:00:00":
            return 0
        m = re.match(r"(\d+)\s+days?,\s+(\d+):(\d+):(\d+)", val)
        if m:
            d, h, mi, s = int(m[1]), int(m[2]), int(m[3]), int(m[4])
            return d * 24 * 60 + h * 60 + mi + round(s / 60)
        m = re.match(r"(\d+):(\d+):(\d+)", val)
        if m:
            return int(m[1]) * 60 + int(m[2]) + round(int(m[3]) / 60)
        # HH:MM 形式（秒なし）
        m = re.match(r"^(\d+):(\d+)$", val)
        if m:
            return int(m[1]) * 60 + int(m[2])
    return 0


def to_hours(val) -> float:
    """互換用: 時間値を浮動小数点hoursに変換（表示用）"""
    return to_minutes(val) / 60


def time_from_val(val):
    """Convert value to datetime.time or None"""
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return None
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        m = re.match(r"(\d+):(\d+)(?::(\d+))?", val)
        if m:
            return time(int(m[1]) % 24, int(m[2]), int(m[3]) if m[3] else 0)
    return None


def early_morning_minutes(row) -> int:
    """6:00〜9:00 の実労働時間（分単位整数）を計算"""
    result = 0
    early_start = time(6, 0)
    early_end = time(9, 0)

    def overlap_min(start_t, end_t) -> int:
        if start_t is None or end_t is None:
            return 0
        s = max(start_t, early_start)
        e = min(end_t, early_end)
        if e <= s:
            return 0
        return (e.hour * 60 + e.minute) - (s.hour * 60 + s.minute)

    for i in range(1, 4):
        s = time_from_val(row.get(f"始業時刻{i}"))
        e = time_from_val(row.get(f"終業時刻{i}"))
        if s is None or e is None:
            continue
        h = overlap_min(s, e)
        # Subtract breaks in this window
        for j in range(1, 4):
            bs = time_from_val(row.get(f"休憩{j}入り時刻"))
            be = time_from_val(row.get(f"休憩{j}戻り時刻"))
            if bs and be:
                h -= overlap_min(bs, be)
        result += max(0, h)

    # Fallback: if no 始業時刻 but 出勤時刻 exists
    if result == 0:
        s = time_from_val(row.get("出勤時刻"))
        e = time_from_val(row.get("退勤時刻"))
        if s and e:
            result = max(0, overlap_min(s, e))

    return result


def early_morning_hours(row) -> float:
    """互換用: 時間単位float"""
    return early_morning_minutes(row) / 60


def _t_to_min(v):
    """time → 分単位int"""
    t = time_from_val(v)
    if t is None:
        return None
    return t.hour * 60 + t.minute


def early_houte_minutes(row, m_houte: int) -> int:
    """法定内残業のうち、6:00-9:00の早朝時間帯に含まれる分数を計算。

    勤務予定開始/退勤時刻と実際の出勤/退勤時刻を比較し、
    スケジュール外で 6:00-9:00 に含まれる勤務時間を算出。
    """
    if m_houte == 0:
        return 0
    early_start = 6 * 60
    early_end = 9 * 60

    pred_start = _t_to_min(row.get("勤務予定開始時刻"))
    pred_end = _t_to_min(row.get("勤務予定退勤時刻"))
    actual_start = _t_to_min(row.get("出勤時刻")) or _t_to_min(row.get("始業時刻1"))
    actual_end = _t_to_min(row.get("退勤時刻")) or _t_to_min(row.get("終業時刻1"))

    if not (pred_start is not None and pred_end is not None
            and actual_start is not None and actual_end is not None):
        return 0
    if pred_end == 0:
        pred_end = 24 * 60
    if actual_end == 0:
        actual_end = 24 * 60

    overlap = 0
    # ① 早朝出勤分: 実出勤 < 予定開始 で、6-9 内
    if actual_start < pred_start:
        s = max(early_start, actual_start)
        e = min(early_end, pred_start)
        if e > s:
            overlap += e - s
    # ② 予定退勤後の残業: 予定退勤 < 実退勤 で、6-9 内
    if pred_end < actual_end:
        s = max(early_start, pred_end)
        e = min(early_end, actual_end)
        if e > s:
            overlap += e - s

    return min(overlap, m_houte)


def yukyu_shinya_minutes(row) -> int:
    """有給日のシフト予定時間のうち深夜時間帯(22:00-翌5:00)に含まれる分数。"""
    s = _t_to_min(row.get("勤務予定開始時刻"))
    e = _t_to_min(row.get("勤務予定退勤時刻"))
    if s is None or e is None:
        return 0
    if e == 0 or e <= s:
        e += 24 * 60  # 日跨ぎ補正

    # 深夜窓: 22:00-29:00 (翌5:00)
    overlap = 0
    for ds, de in [(22 * 60, 29 * 60), (-2 * 60, 5 * 60)]:
        ov = max(0, min(e, de) - max(s, ds))
        overlap += ov
    return overlap


# ─────────────────────────────────────────────
# 従業員マスタ解析
# ─────────────────────────────────────────────
def get_allowances(row: pd.Series):
    """手当を集計して返す (固定手当総額, 割増基礎手当, 控除基礎手当)"""
    total = 0.0
    premium_base = 0.0
    deduct_base = 0.0
    for i in range(1, 11):
        name = row.get(f"手当{i} 名前")
        amount = row.get(f"手当{i} 金額")
        freq = row.get(f"手当{i} 頻度")
        if pd.isna(name) or pd.isna(amount):
            continue
        amount = float(amount)
        freq = "" if pd.isna(freq) else str(freq)
        if freq != "毎月":
            continue
        total += amount
        if str(row.get(f"手当{i} 割増賃金の基礎に含める", "")).strip() == "含める":
            premium_base += amount
        if str(row.get(f"手当{i} 勤怠控除の基礎に含める", "")).strip() == "含める":
            deduct_base += amount
    return total, premium_base, deduct_base


def get_time_wage(row: pd.Series, name_keyword: str) -> float:
    """時間帯ごとの時給から名前でマッチした金額を返す"""
    for i in range(1, 6):
        n = row.get(f"時間帯ごとの時給{i} 名前")
        v = row.get(f"時間帯ごとの時給{i} 金額")
        if pd.notna(n) and pd.notna(v) and name_keyword in str(n):
            return float(v)
    return 0.0


def get_commute_info(row: pd.Series):
    """通勤情報 (単価, 上限, 計算方法) を返す"""
    method = str(row.get("通勤手当の計算方法", "")).strip()
    unit_price = row.get("通勤手当の金額・単価")
    unit_price = float(unit_price) if pd.notna(unit_price) else 0.0
    has_limit = str(row.get("通勤手当の上限額を設定する", "")).strip()
    limit = row.get("通勤手当の上限額")
    limit = float(limit) if pd.notna(limit) else None
    return method, unit_price, has_limit, limit


# ─────────────────────────────────────────────
# 勤怠データ集計
# ─────────────────────────────────────────────
def aggregate_attendance(df_att: pd.DataFrame):
    """
    日次勤怠データを従業員ごとに集計する。
    戻り値: dict[emp_id] = {hours fields, day counts, per_day_records}
    """
    result = {}

    for emp_id, grp in df_att.groupby("従業員番号"):
        # _m サフィックス = 分単位整数で集計（浮動小数点誤差ゼロ）
        agg = {
            "emp_id": emp_id,
            "name": grp["氏名"].iloc[0],
            "dept": grp["部門"].iloc[0],
            "所定内_m": 0, "法定内残業_m": 0, "時間外残業_m": 0,
            "法定休日_m": 0, "深夜_m": 0, "遅刻早退_m": 0,
            "総勤務_m": 0, "早朝_m": 0,
            "早朝法定内残業_m": 0,  # 早朝6-9内の法定内残業
            "深夜有給_m": 0,        # 有給日のシフト予定における深夜時間
            "法定休日越境_m": 0,    # 24時跨ぎで法定休日カウントされた分（修正後は法定内/時間外に移動）
            "法定休日越境_to_houte_m": 0,   # 越境分のうち法定内残業へ移動した分
            "法定休日越境_to_jikan_m": 0,   # 越境分のうち時間外残業へ移動した分
            "出勤日数": 0, "欠勤日数": 0, "実労働日数": 0,
            "per_day": []
        }

        for _, row in grp.iterrows():
            kind = str(row.get("勤怠種別", "")).strip()

            m_teijo = to_minutes(row.get("所定内労働時間"))
            m_houte = to_minutes(row.get("法定内残業時間"))
            m_jiangai = to_minutes(row.get("時間外労働時間"))
            m_hoteiky = to_minutes(row.get("法定休日労働時間"))
            m_shinya = to_minutes(row.get("深夜労働時間"))
            m_late = to_minutes(row.get("控除対象の遅刻早退"))
            m_total = to_minutes(row.get("総勤務時間"))

            row_dict = dict(row)
            m_early = early_morning_minutes(row_dict)
            m_early_houte = early_houte_minutes(row_dict, m_houte)

            YUUKYU = {"有休", "有給特休"}
            # 有給日のシフト予定深夜時間
            m_yukyu_shinya = yukyu_shinya_minutes(row_dict) if kind in YUUKYU else 0

            # 24時跨ぎ法定休日越境の修正
            # 勤務日種別が法定休日でないのに法定休日労働時間がある = 翌日へ越境
            # → 法定休日労働を 0 にし、その日の労働状況に応じて法定内残業 or 時間外残業 に移動
            kinmu_kind = str(row.get("勤務日種別", "")).strip()
            m_cross_to_houte = 0
            m_cross_to_jikan = 0
            if kinmu_kind != "法定休日" and m_hoteiky > 0:
                m_cross = m_hoteiky
                # 法定内 (合計が8h以下) か 時間外 (8h超) かを判定
                total_work_before_cross = m_teijo + m_houte + m_jiangai
                if total_work_before_cross + m_cross <= 8 * 60:
                    m_cross_to_houte = m_cross
                elif total_work_before_cross >= 8 * 60:
                    m_cross_to_jikan = m_cross
                else:
                    # 一部が8h境界をまたぐ場合: 8h以内分は法定内、超過分は時間外
                    to_houte = 8 * 60 - total_work_before_cross
                    m_cross_to_houte = to_houte
                    m_cross_to_jikan = m_cross - to_houte
                # 集計を修正
                m_houte += m_cross_to_houte
                m_jiangai += m_cross_to_jikan
                m_hoteiky = 0
                agg["法定休日越境_m"] += m_cross
                agg["法定休日越境_to_houte_m"] += m_cross_to_houte
                agg["法定休日越境_to_jikan_m"] += m_cross_to_jikan

            if m_total > 0 and kind not in YUUKYU:
                agg["実労働日数"] += 1
            if (m_total > 0 and kind not in YUUKYU) or kind in YUUKYU:
                agg["出勤日数"] += 1
            if kind in ABSENCE_TYPES:
                agg["欠勤日数"] += 1

            agg["所定内_m"] += m_teijo
            agg["法定内残業_m"] += m_houte
            agg["時間外残業_m"] += m_jiangai
            agg["法定休日_m"] += m_hoteiky
            agg["深夜_m"] += m_shinya
            agg["遅刻早退_m"] += m_late
            agg["総勤務_m"] += m_total
            agg["早朝法定内残業_m"] += m_early_houte
            agg["深夜有給_m"] += m_yukyu_shinya
            agg["早朝_m"] += m_early

            try:
                d = pd.to_datetime(row.get("日付")).date()
            except Exception:
                d = None
            weekday = str(row.get("曜日", "")).strip()
            is_taikyo = False
            if d:
                is_taikyo = weekday in ("土", "日") or jpholiday.is_holiday(d)

            # 調整フラグ（日次勤怠シートでハイライト用）
            adjustments = []
            if m_early_houte > 0:
                adjustments.append("早朝法定内残業")
            if m_yukyu_shinya > 0:
                adjustments.append("深夜有給")
            if m_cross_to_houte > 0 or m_cross_to_jikan > 0:
                adjustments.append("法定休日越境")
            if kind in YUUKYU and is_taikyo:
                adjustments.append("土日祝有給")

            agg["per_day"].append({
                "date": d, "is_taikyo": is_taikyo, "kind": kind,
                "所定内_m": m_teijo, "法定内残業_m": m_houte, "時間外残業_m": m_jiangai,
                "法定休日_m": m_hoteiky, "深夜_m": m_shinya,
                "_adjustments": adjustments,
            })

        # 後方互換用: hours単位float (表示用)
        for k in ("所定内", "法定内残業", "時間外残業", "法定休日", "深夜",
                  "遅刻早退", "総勤務", "早朝"):
            agg[k] = agg[f"{k}_m"] / 60

        result[emp_id] = agg

    return result


def aggregate_chitose(per_day_records):
    """千歳フォーマット用に平日/土日祝別に集計（分単位）

    土日祝に取得した有給は、freeeシステム同様に平日扱い（基本支給に計上）。
    """
    YUUKYU = {"有休", "有給特休"}
    heijitsu = {"所定内_m": 0, "法定内残業_m": 0, "時間外残業_m": 0, "法定休日_m": 0, "深夜_m": 0}
    taikyo =   {"所定内_m": 0, "法定内残業_m": 0, "時間外残業_m": 0, "法定休日_m": 0, "深夜_m": 0}
    for day in per_day_records:
        # 土日祝の有給は平日扱い
        is_taikyo_actual = day["is_taikyo"] and day.get("kind") not in YUUKYU
        target = taikyo if is_taikyo_actual else heijitsu
        for k in target:
            target[k] += day.get(k, 0)
    return heijitsu, taikyo


# ─────────────────────────────────────────────
# 給与計算
# ─────────────────────────────────────────────
def calc_commute(method, unit_price, has_limit, limit, jissai_days) -> int:
    if not method or unit_price == 0:
        return 0
    if "日数" in method:
        amount = unit_price * jissai_days
    else:
        amount = unit_price  # 直接入力 or 月額
    if has_limit == "設定する" and limit:
        amount = min(amount, limit)
    return ceil_int(amount)


def payroll_standard(agg: dict, emp: pd.Series, has_hayao: bool) -> dict:
    """標準フォーマット給与計算 (旭川・すすきの・おはこ)"""
    base_pay = float(emp.get("基本給", 0) or 0)
    wage_type = str(emp.get("給与方式", "")).strip()
    fixed_total, premium_base, deduct_base = get_allowances(emp)
    hayao_wage = get_time_wage(emp, "早朝") if has_hayao else 0.0

    method, unit_price, has_limit, limit = get_commute_info(emp)
    jissai = agg["実労働日数"]

    # 分単位整数で計算
    m_teijo = agg["所定内_m"]
    m_houte = agg["法定内残業_m"]
    m_jikan = agg["時間外残業_m"]
    m_hoteiky = agg["法定休日_m"]
    m_shinya = agg["深夜_m"]
    m_late = agg["遅刻早退_m"]
    m_early = agg["早朝_m"] if has_hayao else 0
    m_early_houte = agg["早朝法定内残業_m"] if has_hayao else 0  # 早朝6-9内の法定内残業
    m_yukyu_shinya = agg["深夜有給_m"]  # 有給の予定シフト中の深夜時間
    kekkin_days = agg["欠勤日数"]
    # 表示用float
    h_teijo = m_teijo / 60
    h_houte = m_houte / 60
    h_jikan = m_jikan / 60
    h_hoteiky = m_hoteiky / 60
    h_shinya = m_shinya / 60
    h_late = m_late / 60
    h_early = m_early / 60
    h_early_houte = m_early_houte / 60

    # 早朝法定内残業以外の法定内残業 (= 基本時給で計算する分)
    m_houte_normal = m_houte - m_early_houte

    if wage_type == "月給":
        premium_rate = (base_pay + premium_base) / MONTHLY_STD_HOURS
        deduct_rate = (base_pay + deduct_base) / MONTHLY_STD_HOURS
        kihon_shikyuu = ceil_int(base_pay)
        jikanwari = 0.0
        hayao_rate = 0.0
        houte_pay = 0  # 月給は法定内残業手当なし
        houte_early_pay = 0
        jikan_pay = amount_by_min(premium_rate, m_jikan, 1.25, "ceil")
        shinya_pay = amount_by_min(premium_rate, m_shinya, 0.25, "ceil")
        hoteiky_pay = amount_by_min(premium_rate, m_hoteiky, 1.35, "ceil")
        hayao_pay = 0
        # 深夜有給割増は月給者は対象外
        shinya_yukyu_pay = 0
        m_yukyu_shinya = 0  # 表示も0に
        late_ded = amount_by_min(deduct_rate, m_late, 1.0, "floor")
        kekkin_ded = floor_int(deduct_rate * kekkin_days * 8)
    else:
        jikanwari = base_pay
        premium_rate = jikanwari
        deduct_rate = jikanwari
        hayao_rate = hayao_wage if hayao_wage else ceil_int(jikanwari * 1.25)
        # 基本支給 = 時給 × (所定内 - 早朝所定内分) / 60
        # 早朝所定内分 = m_early - m_early_houte
        m_early_teijo = m_early - m_early_houte
        kihon_shikyuu = amount_by_min(jikanwari, m_teijo - m_early_teijo, 1.0, "ceil")
        # 法定内残業手当（早朝以外）= 基本時給 × (法定内残業 - 早朝法定内残業) / 60
        houte_pay = amount_by_min(jikanwari, m_houte_normal, 1.0, "ceil")
        # 法定内残業手当（早朝分）= 早朝時給 × 早朝法定内残業 / 60
        houte_early_pay = amount_by_min(hayao_rate, m_early_houte, 1.0, "ceil") if has_hayao else 0
        jikan_pay = amount_by_min(jikanwari, m_jikan, 1.25, "ceil")
        shinya_pay = amount_by_min(jikanwari, m_shinya, 0.25, "ceil")
        hoteiky_pay = amount_by_min(jikanwari, m_hoteiky, 1.35, "ceil")
        # 早朝手当 = 早朝時給 × 早朝所定内分 / 60 (早朝法定内残業は別計上)
        hayao_pay = amount_by_min(hayao_rate, m_early_teijo, 1.0, "ceil") if has_hayao else 0
        # 深夜有給割増 = 基本時給 × 深夜有給時間 × 0.25
        shinya_yukyu_pay = amount_by_min(jikanwari, m_yukyu_shinya, 0.25, "ceil")
        late_ded = 0
        kekkin_ded = 0

    commute = calc_commute(method, unit_price, has_limit, limit, jissai)

    # 深夜割増 = 通常深夜割増 + 深夜有給割増
    shinya_pay_total = shinya_pay + shinya_yukyu_pay
    # 法定内残業手当 合計 = 通常分（基本時給）+ 早朝分（早朝時給）
    houte_pay_total = houte_pay + houte_early_pay

    total = (kihon_shikyuu + houte_pay_total + int(fixed_total) + jikan_pay + shinya_pay_total
             + hoteiky_pay + hayao_pay - late_ded - kekkin_ded + commute)

    # ハイライト対象
    highlights = []
    if m_yukyu_shinya > 0:
        highlights.append("深夜割増(0.25)")
    # 早朝法定内残業がある場合は両方の金額列をハイライト
    if has_hayao and m_early_houte > 0:
        highlights.extend(["早朝法定内残業手当", "法定内残業手当_合計"])
    # 法定休日越境調整があった場合: 法定休日(0)・移動先をハイライト
    m_cross = agg.get("法定休日越境_m", 0)
    if m_cross > 0:
        highlights.append("法定休日")
        highlights.append("法定休日手当(1.35)")
        if agg.get("法定休日越境_to_houte_m", 0) > 0:
            highlights.append("法定内残業")
            highlights.append("法定内残業手当")
            highlights.append("法定内残業手当_合計")
        if agg.get("法定休日越境_to_jikan_m", 0) > 0:
            highlights.append("時間外残業")
            highlights.append("時間外手当(1.25)")

    # freee表示順: 基本給→[早朝]→勤怠控除→残業手当合計→手当合計
    row: dict = {
        "従業員番号": agg["emp_id"],
        "氏名": agg["name"],
        "給与方式": wage_type,
        "基本給/時給": base_pay if wage_type == "月給" else jikanwari,
        "出勤日数": agg["出勤日数"],
        "欠勤日数": kekkin_days,
        "実労働日数": jissai,
        # ─ 基本給ブロック ─
        "所定内": h_teijo,
        "基本支給額": kihon_shikyuu,
    }
    # ─ 早朝ブロック ─
    if has_hayao:
        row["早朝6-9"] = h_early
        row["早朝時給"] = hayao_rate
        row["早朝手当"] = hayao_pay
    # ─ 勤怠控除ブロック ─
    row.update({
        "遅刻早退": h_late,
        "遅刻早退控除": late_ded,
        "欠勤控除": kekkin_ded,
    })
    # ─ 残業手当合計ブロック ─（時間→金額の順）
    row["法定内残業"] = h_houte
    row["法定内残業手当"] = houte_pay  # 早朝以外（基本時給で計算）
    if has_hayao:
        row["早朝法定内残業"] = h_early_houte
        row["早朝法定内残業手当"] = houte_early_pay  # 早朝時給で計算
        row["法定内残業手当_合計"] = houte_pay_total
    row.update({
        "時間外残業": h_jikan,
        "時間外手当(1.25)": jikan_pay,
        "法定休日": h_hoteiky,
        "法定休日手当(1.35)": hoteiky_pay,
        "深夜": h_shinya,
        "深夜割増(0.25)": shinya_pay_total,
        "深夜有給": m_yukyu_shinya / 60,
        "深夜有給割増": shinya_yukyu_pay,
    })
    # ─ 手当合計・単価ブロック ─
    row.update({
        "固定手当(総)": int(fixed_total),
        "割増基礎手当": int(premium_base),
        "控除基礎手当": int(deduct_base),
        "割増単価": round(premium_rate, 4) if wage_type == "月給" else premium_rate,
        "控除単価": round(deduct_rate, 4) if wage_type == "月給" else deduct_rate,
    })
    # ─ 通勤・総支給ブロック ─
    row.update({
        "通勤単価": unit_price,
        "通勤上限": int(limit) if limit else "",
        "通勤手当": commute,
        "総勤務": agg["総勤務"],
        "総支給額": total,
        "_highlights": highlights,
        "_per_day_adjustments": agg["per_day"],
    })
    return row


def payroll_chitose(agg: dict, emp: pd.Series) -> dict:
    """千歳フォーマット給与計算 (土日祝分離)"""
    base_pay = float(emp.get("基本給", 0) or 0)
    wage_type = str(emp.get("給与方式", "")).strip()
    fixed_total, premium_base, deduct_base = get_allowances(emp)
    taikyo_wage = get_time_wage(emp, "土日祝")

    method, unit_price, has_limit, limit = get_commute_info(emp)
    jissai = agg["実労働日数"]
    kekkin_days = agg["欠勤日数"]

    heijitsu, taikyo = aggregate_chitose(agg["per_day"])

    # 分単位整数で集計
    m_teijo_all = agg["所定内_m"]
    m_houte_all = agg["法定内残業_m"]
    m_jikan_all = agg["時間外残業_m"]
    m_hoteiky_all = agg["法定休日_m"]
    m_shinya_all = agg["深夜_m"]
    m_late = agg["遅刻早退_m"]

    m_taikyo_teijo = taikyo["所定内_m"]
    m_taikyo_houte = taikyo["法定内残業_m"]
    m_taikyo_jikan = taikyo["時間外残業_m"]
    m_taikyo_hoteiky = taikyo["法定休日_m"]
    m_taikyo_shinya = taikyo["深夜_m"]

    m_hei_jikan = heijitsu["時間外残業_m"]
    m_hei_hoteiky = heijitsu["法定休日_m"]
    m_hei_shinya = heijitsu["深夜_m"]
    m_hei_houte = heijitsu["法定内残業_m"]
    m_yukyu_shinya = agg["深夜有給_m"]  # 有給日のシフト予定深夜時間

    # 表示用float
    h_teijo_all = m_teijo_all / 60
    h_houte_all = m_houte_all / 60
    h_jikan_all = m_jikan_all / 60
    h_hoteiky_all = m_hoteiky_all / 60
    h_shinya_all = m_shinya_all / 60
    h_late = m_late / 60
    h_taikyo_teijo = m_taikyo_teijo / 60
    h_taikyo_houte = m_taikyo_houte / 60
    h_taikyo_jikan = m_taikyo_jikan / 60
    h_taikyo_hoteiky = m_taikyo_hoteiky / 60
    h_taikyo_shinya = m_taikyo_shinya / 60

    if wage_type == "月給":
        premium_rate = (base_pay + premium_base) / MONTHLY_STD_HOURS
        deduct_rate = (base_pay + deduct_base) / MONTHLY_STD_HOURS
        kihon_shikyuu = ceil_int(base_pay)
        taikyo_kihon = 0
        # 月給は割増単価が一定なので、全て平日列に計上（参照シートに合わせる）
        jikan_hei = amount_by_min(premium_rate, m_jikan_all, 1.25, "ceil")
        jikan_taikyo = 0
        shinya_hei = amount_by_min(premium_rate, m_shinya_all, 0.25, "ceil")
        shinya_taikyo = 0
        hoteiky_hei = amount_by_min(premium_rate, m_hoteiky_all, 1.35, "ceil")
        hoteiky_taikyo = 0
        houte_hei = amount_by_min(premium_rate, m_houte_all, 0.25, "ceil")
        houte_taikyo = 0
        # 深夜有給割増は月給者は対象外
        shinya_yukyu_pay = 0
        m_yukyu_shinya = 0
        late_ded = amount_by_min(deduct_rate, m_late, 1.0, "floor")
        kekkin_ded = floor_int(deduct_rate * kekkin_days * 8)
    else:
        jikanwari = base_pay
        premium_rate = jikanwari
        deduct_rate = jikanwari
        # 千歳の慣行: 土日祝時給が未設定の場合は基本時給+100円
        taikyo_rate = taikyo_wage if taikyo_wage else (jikanwari + 100)

        # 基本支給: 平日所定内 × 基本時給、土日祝基本: 土日祝所定 × 土日祝時給
        m_hei_teijo = m_teijo_all - m_taikyo_teijo
        kihon_shikyuu = amount_by_min(jikanwari, m_hei_teijo, 1.0, "ceil")
        taikyo_kihon = amount_by_min(taikyo_rate, m_taikyo_teijo, 1.0, "ceil")

        # 時間外/深夜/法定休日/法定内残業: 差額方式
        #   平日列 = 基本時給 × 全時間 × 割増率
        #   土日祝列 = (土日祝時給 - 基本時給) × 土日祝時間 × 割増率
        diff_rate = taikyo_rate - jikanwari
        jikan_hei = amount_by_min(jikanwari, m_jikan_all, 1.25, "ceil")
        jikan_taikyo = amount_by_min(diff_rate, m_taikyo_jikan, 1.25, "ceil")
        shinya_hei = amount_by_min(jikanwari, m_shinya_all, 0.25, "ceil")
        shinya_taikyo = amount_by_min(diff_rate, m_taikyo_shinya, 0.25, "ceil")
        hoteiky_hei = amount_by_min(jikanwari, m_hoteiky_all, 1.35, "ceil")
        hoteiky_taikyo = amount_by_min(diff_rate, m_taikyo_hoteiky, 1.35, "ceil")
        houte_hei = amount_by_min(jikanwari, m_houte_all, 1.0, "ceil")
        houte_taikyo = amount_by_min(diff_rate, m_taikyo_houte, 1.0, "ceil")
        # 深夜有給割増 = 基本時給 × 深夜有給時間 × 0.25
        shinya_yukyu_pay = amount_by_min(jikanwari, m_yukyu_shinya, 0.25, "ceil")
        late_ded = 0
        kekkin_ded = 0

    commute = calc_commute(method, unit_price, has_limit, limit, jissai)

    # 深夜割増_平日に深夜有給割増を加算
    shinya_hei_total = shinya_hei + shinya_yukyu_pay

    total = (kihon_shikyuu + taikyo_kihon + int(fixed_total)
             + jikan_hei + jikan_taikyo
             + shinya_hei_total + shinya_taikyo
             + hoteiky_hei + hoteiky_taikyo
             + houte_hei + houte_taikyo
             - late_ded - kekkin_ded + commute)

    # ハイライト対象セル: 時給者のみ、土日祝関連を要確認
    highlights = []
    if wage_type == "時給":
        has_taikyo_yukyu = any(
            d.get("kind") in {"有休", "有給特休"} and d.get("is_taikyo") and d.get("所定内_m", 0) > 0
            for d in agg["per_day"]
        )
        if has_taikyo_yukyu:
            highlights.extend(["基本支給額", "土日祝基本"])
        if m_taikyo_jikan > 0:
            highlights.append("時間外手当_合計")
        if m_taikyo_shinya > 0:
            highlights.append("深夜割増_合計")
        if m_taikyo_hoteiky > 0:
            highlights.append("法定休日手当_合計")
        if m_taikyo_houte > 0:
            highlights.append("法定内残業_合計")
    # 深夜有給がある場合は深夜割増_合計をハイライト
    if m_yukyu_shinya > 0:
        if "深夜割増_合計" not in highlights:
            highlights.append("深夜割増_合計")
    # 法定休日越境調整があった場合
    m_cross = agg.get("法定休日越境_m", 0)
    if m_cross > 0:
        highlights.append("法定休日")
        highlights.append("法定休日手当_平日")
        highlights.append("法定休日手当_合計")
        if agg.get("法定休日越境_to_houte_m", 0) > 0:
            highlights.append("法定内残業")
            highlights.append("法定内残業_平日")
            highlights.append("法定内残業_合計")
        if agg.get("法定休日越境_to_jikan_m", 0) > 0:
            highlights.append("時間外残業")
            highlights.append("時間外手当_平日")
            highlights.append("時間外手当_合計")

    # freee表示順: 基本給→勤怠控除→残業手当合計→手当合計
    row = {
        "従業員番号": agg["emp_id"],
        "氏名": agg["name"],
        "給与方式": wage_type,
        "基本給/時給": base_pay,
        "出勤日数": agg["出勤日数"],
        "欠勤日数": kekkin_days,
        "実労働日数": jissai,
        # ─ 基本給ブロック ─
        "所定内": h_teijo_all,
        "基本支給額": kihon_shikyuu,
        # ─ 土日祝基本ブロック ─
        "土日祝時給": taikyo_rate if wage_type != "月給" else "",
        "土日祝所定": h_taikyo_teijo,
        "土日祝基本": taikyo_kihon,
        # ─ 勤怠控除ブロック ─
        "遅刻早退": h_late,
        "遅刻早退控除": late_ded,
        "欠勤控除": kekkin_ded,
        # ─ 残業手当合計ブロック ─（種類ごとに 時間→金額 をまとめる）
        # 法定内残業
        "法定内残業": h_houte_all,
        "法定内残業_平日": houte_hei,
        "土日祝法定内": h_taikyo_houte,
        "法定内残業_土日祝": houte_taikyo,
        "法定内残業_合計": houte_hei + houte_taikyo,
        # 時間外残業
        "時間外残業": h_jikan_all,
        "時間外手当_平日": jikan_hei,
        "土日祝時間外": h_taikyo_jikan,
        "時間外手当_土日祝": jikan_taikyo,
        "時間外手当_合計": jikan_hei + jikan_taikyo,
        # 法定休日
        "法定休日": h_hoteiky_all,
        "法定休日手当_平日": hoteiky_hei,
        "土日祝法定休日": h_taikyo_hoteiky,
        "法定休日手当_土日祝": hoteiky_taikyo,
        "法定休日手当_合計": hoteiky_hei + hoteiky_taikyo,
        # 深夜
        "深夜": h_shinya_all,
        "深夜割増_平日": shinya_hei_total,
        "土日祝深夜": h_taikyo_shinya,
        "深夜割増_土日祝": shinya_taikyo,
        "深夜割増_合計": shinya_hei_total + shinya_taikyo,
        "深夜有給": m_yukyu_shinya / 60 if m_yukyu_shinya > 0 else 0,
        "深夜有給割増": shinya_yukyu_pay if m_yukyu_shinya > 0 else 0,
        # ─ 手当合計・単価ブロック ─
        "固定手当(総)": int(fixed_total),
        "割増基礎手当": int(premium_base),
        "controlling_控除基礎手当": int(deduct_base),
        "割増単価": round(premium_rate, 4) if wage_type == "月給" else premium_rate,
        "控除単価": round(deduct_rate, 4) if wage_type == "月給" else deduct_rate,
        # ─ 通勤・総支給ブロック ─
        "通勤単価": unit_price,
        "通勤上限": int(limit) if limit else "",
        "通勤手当": commute,
        "総勤務": agg["総勤務"],
        "総支給額": total,
        "_highlights": highlights,
        "_per_day_adjustments": agg["per_day"],
    }
    # キー名のtypo修正
    row["控除基礎手当"] = row.pop("controlling_控除基礎手当")
    # 順序を維持するため、再構築
    ordered_keys = [
        "従業員番号", "氏名", "給与方式", "基本給/時給",
        "出勤日数", "欠勤日数", "実労働日数",
        "所定内", "基本支給額",
        "土日祝時給", "土日祝所定", "土日祝基本",
        "遅刻早退", "遅刻早退控除", "欠勤控除",
        "法定内残業", "法定内残業_平日", "土日祝法定内", "法定内残業_土日祝", "法定内残業_合計",
        "時間外残業", "時間外手当_平日", "土日祝時間外", "時間外手当_土日祝", "時間外手当_合計",
        "法定休日", "法定休日手当_平日", "土日祝法定休日", "法定休日手当_土日祝", "法定休日手当_合計",
        "深夜", "深夜割増_平日", "土日祝深夜", "深夜割増_土日祝", "深夜割増_合計",
        "深夜有給", "深夜有給割増",
        "固定手当(総)", "割増基礎手当", "控除基礎手当", "割増単価", "控除単価",
        "通勤単価", "通勤上限", "通勤手当",
        "総勤務", "総支給額",
        "_highlights", "_per_day_adjustments",
    ]
    return {k: row[k] for k in ordered_keys}


# ─────────────────────────────────────────────
# Excel出力
# ─────────────────────────────────────────────
TIME_COLS = {"所定内", "法定内残業", "時間外残業", "法定休日", "深夜", "遅刻早退",
             "総勤務", "早朝6-9", "早朝法定内残業", "深夜有給",
             "土日祝所定", "土日祝法定内", "土日祝時間外", "土日祝法定休日", "土日祝深夜"}


def fmt_hours(h: float) -> str:
    total_min = round(h * 60)
    sign = "-" if total_min < 0 else ""
    total_min = abs(total_min)
    hrs = total_min // 60
    mins = total_min % 60
    return f"{sign}{hrs}:{mins:02d}"


def style_header(cell):
    cell.font = Font(name="Arial", bold=True, size=10)
    cell.fill = PatternFill("solid", start_color="D9D9D9")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_data(cell, is_subtotal=False):
    cell.font = Font(name="Arial", size=10, bold=is_subtotal)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if is_subtotal:
        cell.fill = PatternFill("solid", start_color="FFFFC0")


def write_store_sheet(wb, store_name: str, rows: list[dict], store_fmt: str,
                      period_str: str, std_hours: float):
    ws = wb.create_sheet(title=store_name[:31])

    # タイトル
    title = f"給与計算 — {store_name}（{period_str}）"
    ws.cell(1, 1, title).font = Font(name="Arial", bold=True, size=12)
    ws.cell(2, 1, f"月平均所定労働時間：{std_hours}h").font = Font(name="Arial", size=10)

    if not rows:
        return

    # _highlights は内部用なので表示列から除外
    headers = [k for k in rows[0].keys() if not k.startswith("_")]
    header_row = 4

    # ヘッダ
    for c, h in enumerate(headers, 1):
        cell = ws.cell(header_row, c, h)
        style_header(cell)
        ws.column_dimensions[get_column_letter(c)].width = max(10, min(len(h) * 2, 20))

    # 土日祝関連ハイライト用 (オレンジ系)
    HIGHLIGHT_FILL = PatternFill("solid", start_color="FFD699")

    # データ
    for r_idx, row_data in enumerate(rows):
        excel_row = header_row + 1 + r_idx
        highlights = set(row_data.get("_highlights", []))
        for c_idx, key in enumerate(headers, 1):
            val = row_data[key]
            if key in TIME_COLS and isinstance(val, float):
                val = fmt_hours(val)
            cell = ws.cell(excel_row, c_idx, val)
            style_data(cell)
            if key in highlights:
                cell.fill = HIGHLIGHT_FILL

    # 合計行
    sum_row = header_row + 1 + len(rows)
    SUM_INT_COLS = {"出勤日数", "欠勤日数", "実労働日数",
                    "固定手当(総)", "割増基礎手当", "控除基礎手当",
                    "基本支給額", "時間外手当(1.25)", "深夜割増(0.25)",
                    "法定休日手当(1.35)", "早朝手当",
                    "法定内残業手当", "早朝法定内残業手当", "法定内残業手当_合計",
                    "深夜有給割増",
                    "遅刻早退控除", "欠勤控除", "通勤手当", "総支給額",
                    "土日祝基本",
                    "時間外手当_平日", "時間外手当_土日祝", "時間外手当_合計",
                    "深夜割増_平日", "深夜割増_土日祝", "深夜割増_合計",
                    "法定休日手当_平日", "法定休日手当_土日祝", "法定休日手当_合計",
                    "法定内残業_平日", "法定内残業_土日祝", "法定内残業_合計",
                    }

    for c_idx, key in enumerate(headers, 1):
        if key == "従業員番号":
            cell = ws.cell(sum_row, c_idx, "合計")
        elif key in SUM_INT_COLS:
            total = sum(r[key] for r in rows if isinstance(r.get(key), (int, float)))
            cell = ws.cell(sum_row, c_idx, round(total))
        elif key in TIME_COLS:
            total_h = sum(r[key] for r in rows if isinstance(r.get(key), float))
            cell = ws.cell(sum_row, c_idx, fmt_hours(total_h))
        else:
            cell = ws.cell(sum_row, c_idx, "")
        style_data(cell, is_subtotal=True)

    # 列幅調整
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14


def write_attendance_sheet(wb, df_att: pd.DataFrame, store_results: dict):
    """元の日次勤怠データを記入し、調整対象日に色付け"""
    # 各従業員×日付の調整種別を集約
    # key: (emp_id_int, date_str), value: set of adjustment labels
    adj_map: dict = {}
    for store, (rows, _fmt) in store_results.items():
        for r in rows:
            eid = int(r["従業員番号"])
            # _per_dayがrowに保存されていないので、aggから取得する必要がある
            # → rowに保存しておくよう変更が必要
            for day in r.get("_per_day_adjustments", []):
                d = day["date"]
                if not d or not day.get("_adjustments"):
                    continue
                key = (eid, str(d))
                adj_map.setdefault(key, set()).update(day["_adjustments"])

    ws = wb.create_sheet(title="日次勤怠（元データ）", index=0)
    ws.cell(1, 1, "日次勤怠一覧（処理対象日に色付け）").font = Font(name="Arial", bold=True, size=12)
    ws.cell(2, 1, "凡例: 早朝法定内残業=黄, 深夜有給=水色, 法定休日越境=橙, 土日祝有給=ピンク").font = Font(name="Arial", size=9)

    headers = list(df_att.columns)
    header_row = 4
    for c, h in enumerate(headers, 1):
        cell = ws.cell(header_row, c, str(h))
        style_header(cell)
        ws.column_dimensions[get_column_letter(c)].width = max(10, min(len(str(h)) * 1.5, 22))

    # 色定義
    FILL_HAYAO = PatternFill("solid", start_color="FFF2CC")   # 黄
    FILL_YUKYU_SHINYA = PatternFill("solid", start_color="DAEEF3")  # 水色
    FILL_CROSS = PatternFill("solid", start_color="FFD699")    # 橙
    FILL_TAIKYO_YUKYU = PatternFill("solid", start_color="FCE4D6")  # ピンク
    FILL_MAP = {
        "早朝法定内残業": FILL_HAYAO,
        "深夜有給": FILL_YUKYU_SHINYA,
        "法定休日越境": FILL_CROSS,
        "土日祝有給": FILL_TAIKYO_YUKYU,
    }
    # 優先順位（複数該当時は法定休日越境を優先）
    PRIORITY = ["法定休日越境", "深夜有給", "早朝法定内残業", "土日祝有給"]

    # データ行
    for r_idx, (_, row) in enumerate(df_att.iterrows()):
        excel_row = header_row + 1 + r_idx
        # 調整種別取得
        try:
            eid = int(pd.to_numeric(row.get("従業員番号"), errors="coerce"))
        except (ValueError, TypeError):
            eid = None
        try:
            d_val = pd.to_datetime(row.get("日付")).date()
            d_str = str(d_val)
        except Exception:
            d_str = None
        adjustments = adj_map.get((eid, d_str), set()) if eid and d_str else set()
        fill = None
        for label in PRIORITY:
            if label in adjustments:
                fill = FILL_MAP[label]
                break

        for c_idx, key in enumerate(headers, 1):
            val = row[key]
            if pd.isna(val):
                val = None
            elif isinstance(val, pd.Timestamp):
                val = val.strftime("%Y-%m-%d")
            elif isinstance(val, time):
                val = val.strftime("%H:%M:%S")
            elif isinstance(val, timedelta):
                total_sec = int(val.total_seconds())
                val = f"{total_sec//3600}:{(total_sec%3600)//60:02d}:{total_sec%60:02d}"
            cell = ws.cell(excel_row, c_idx, val)
            cell.font = Font(name="Arial", size=9)
            if fill:
                cell.fill = fill

    ws.freeze_panes = "C5"


def create_excel(store_results: dict, period_str: str, std_hours: float,
                 df_att: pd.DataFrame = None) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for store_name, (rows, fmt) in store_results.items():
        write_store_sheet(wb, store_name, rows, fmt, period_str, std_hours)

    # 元の日次勤怠シートを追加（最初のシートに）
    if df_att is not None:
        write_attendance_sheet(wb, df_att, store_results)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────
# おはこ専用処理 (月集計CSV)
# ─────────────────────────────────────────────
def parse_days(val) -> int:
    """'23 日' → 23 のように数値抽出"""
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return 0
    s = str(val).strip()
    m = re.match(r"(\d+)", s)
    return int(m[1]) if m else 0


def process_ohako(df_book7: pd.DataFrame, df_emp: pd.DataFrame,
                  std_hours: float, store_name: str = "焼肉ホルモン　おはこ忠和店") -> dict:
    """おはこ用: 月集計CSVを処理（per-day情報なし、早朝/越境/深夜有給は非対応）"""
    global MONTHLY_STD_HOURS
    MONTHLY_STD_HOURS = std_hours

    df_book7 = df_book7.copy()
    df_book7["従業員番号"] = pd.to_numeric(df_book7["従業員番号"], errors="coerce")
    df_book7 = df_book7[df_book7["従業員番号"].notna()].copy()
    df_book7["従業員番号"] = df_book7["従業員番号"].astype(int)

    df_emp = df_emp.copy()
    df_emp["従業員番号"] = pd.to_numeric(df_emp["従業員番号"], errors="coerce")
    df_emp["_eid"] = df_emp["従業員番号"].apply(lambda x: float(x) if pd.notna(x) else None)

    rows = []
    for _, brow in df_book7.iterrows():
        eid = int(brow["従業員番号"])
        emp_rows = df_emp[df_emp["_eid"] == float(eid)]
        if emp_rows.empty:
            continue
        emp = emp_rows.iloc[0]

        # 月集計データを内部agg形式に変換
        # 「法定内」= 所定内 + 法定内残業 の合計（時給者にとっては支払い単価が同じなので所定として扱う）
        m_houtei = to_minutes(brow.get("法定内"))
        agg = {
            "emp_id": eid,
            "name": str(emp.get("freee人事労務での表示名") or brow.get("氏名") or ""),
            "dept": str(brow.get("部門", "")),
            "所定内_m": m_houtei,
            "法定内残業_m": 0,
            "時間外残業_m": to_minutes(brow.get("時間外")),
            "法定休日_m": to_minutes(brow.get("法定")),
            "深夜_m": to_minutes(brow.get("深夜")),
            "遅刻早退_m": to_minutes(brow.get("遅刻")),
            "総勤務_m": to_minutes(brow.get("総勤務")),
            "早朝_m": 0,
            "早朝法定内残業_m": 0,
            "深夜有給_m": 0,
            "法定休日越境_m": 0,
            "法定休日越境_to_houte_m": 0,
            "法定休日越境_to_jikan_m": 0,
            "出勤日数": parse_days(brow.get("労働")),
            "欠勤日数": parse_days(brow.get("欠勤")),
            "実労働日数": max(0, parse_days(brow.get("労働")) - parse_days(brow.get("欠勤"))),
            "per_day": [],
        }
        for k in ("所定内", "法定内残業", "時間外残業", "法定休日", "深夜",
                  "遅刻早退", "総勤務", "早朝"):
            agg[k] = agg[f"{k}_m"] / 60

        # おはこは早朝時給なし
        row_data = payroll_standard(agg, emp, has_hayao=False)
        rows.append(row_data)

    return {store_name: (rows, "standard")}


# ─────────────────────────────────────────────
# メイン処理
# ─────────────────────────────────────────────
def detect_period(df_att: pd.DataFrame) -> str:
    try:
        dates = pd.to_datetime(df_att["日付"], errors="coerce").dropna()
        if len(dates) == 0:
            return "不明"
        mn, mx = dates.min(), dates.max()
        return f"{mn.year}年{mn.month}月"
    except Exception:
        return "不明"


def process(df_att: pd.DataFrame, df_emp: pd.DataFrame, std_hours: float):
    global MONTHLY_STD_HOURS
    MONTHLY_STD_HOURS = std_hours

    df_emp = df_emp.copy()
    df_emp["従業員番号"] = pd.to_numeric(df_emp["従業員番号"], errors="coerce")
    df_emp["_eid"] = df_emp["従業員番号"].apply(lambda x: int(x) if pd.notna(x) else None)

    # 部門1でストア分類 (int key)
    emp_store_map = {}
    for _, row in df_emp.iterrows():
        eid = row.get("_eid")
        store = str(row.get("部門1", "")).strip()
        if eid is not None and store:
            emp_store_map[eid] = store

    # 勤怠集計
    df_att = df_att.copy()
    df_att["従業員番号"] = pd.to_numeric(df_att["従業員番号"], errors="coerce")
    df_att = df_att[df_att["従業員番号"].notna()]
    df_att["従業員番号"] = df_att["従業員番号"].astype(int)

    att_agg = aggregate_attendance(df_att)

    # ストアごとに集計
    store_results = {}
    skip_keywords = {"本社", "管理部", "総務部"}

    # ─ パス1: 各ストアで早朝時給を持つ従業員がいるか事前チェック ─
    store_has_hayao: dict[str, bool] = {}
    for emp_id, agg in att_agg.items():
        store = emp_store_map.get(int(emp_id), "")
        if not store or any(kw in store for kw in skip_keywords):
            continue
        if store in store_has_hayao:
            continue
        emp_rows = df_emp[df_emp["_eid"] == int(emp_id)]
        if emp_rows.empty:
            continue
        # 店舗の全従業員を一度にチェック
        store_emps = df_emp[df_emp["部門1"] == store]
        store_has_hayao[store] = any(
            get_time_wage(row, "早朝") > 0
            for _, row in store_emps.iterrows()
        )

    # ─ パス2: 各従業員の給与計算 ─
    for emp_id, agg in att_agg.items():
        store = emp_store_map.get(int(emp_id), "")
        if not store or any(kw in store for kw in skip_keywords):
            continue

        emp_rows = df_emp[df_emp["_eid"] == int(emp_id)]
        if emp_rows.empty:
            continue
        emp = emp_rows.iloc[0]

        fmt = STORE_FORMAT.get(store, "standard")
        # 店舗に早朝時給者が1人でもいれば全員に早朝列を出す
        has_hayao = store_has_hayao.get(store, False)

        if fmt == "chitose":
            row_data = payroll_chitose(agg, emp)
        else:
            row_data = payroll_standard(agg, emp, has_hayao)

        if store not in store_results:
            store_results[store] = ([], fmt)
        store_results[store][0].append(row_data)

    return store_results


# ─────────────────────────────────────────────
# Streamlit UI
# ─────────────────────────────────────────────
st.set_page_config(page_title="給与計算アプリ", page_icon="💴", layout="wide")
st.title("💴 給与計算アプリ（freee連携）")
st.caption("日次勤怠CSV・従業員エクスポートCSVをアップロードして給与計算Excelを生成します。")

with st.sidebar:
    st.header("⚙️ 設定")
    std_hours = st.number_input(
        "月平均所定労働時間",
        value=MONTHLY_STD_HOURS, step=0.01, format="%.2f",
        help="基本給÷この時間数で1時間単価を計算します"
    )
    st.markdown("---")
    st.markdown("**ストアフォーマット**")
    st.caption("千歳北信濃店のみ土日祝分離計算、旭川買物公園通り店は早朝時給対応")

tab_komeda, tab_ohako = st.tabs(["☕ コメダ（日次勤怠）", "🥩 おはこ（月集計）"])


def _read_csv_auto(file):
    try:
        return pd.read_csv(file, encoding="utf-8-sig")
    except UnicodeDecodeError:
        file.seek(0)
        return pd.read_csv(file, encoding="shift_jis")


# ─── コメダタブ ─────────────────────────────
with tab_komeda:
    col1, col2 = st.columns(2)
    with col1:
        st.subheader("📋 日次勤怠一覧 CSV")
        att_file = st.file_uploader("日次勤怠一覧（Sheet1）をCSVでアップロード", type=["csv"], key="att")
    with col2:
        st.subheader("👥 従業員エクスポート CSV")
        emp_file = st.file_uploader("employee_exports（Sheet2）をCSVでアップロード", type=["csv"], key="emp")

    if att_file and emp_file:
        df_att = _read_csv_auto(att_file)
        df_emp = _read_csv_auto(emp_file)

        with st.expander("📊 データプレビュー"):
            st.write(f"**勤怠データ** — {len(df_att):,} 行 × {len(df_att.columns)} 列")
            st.dataframe(df_att.head(5), use_container_width=True)
            st.write(f"**従業員データ** — {len(df_emp):,} 行 × {len(df_emp.columns)} 列")
            st.dataframe(df_emp.head(5), use_container_width=True)

        period_str = detect_period(df_att)
        st.info(f"対象期間: **{period_str}**")

        if st.button("🔄 給与計算を実行", type="primary", key="run_komeda"):
            with st.spinner("計算中..."):
                try:
                    store_results = process(df_att, df_emp, std_hours)
                except Exception as e:
                    st.error(f"計算エラー: {e}")
                    st.exception(e)
                    st.stop()

            if not store_results:
                st.warning("計算結果が空です。データを確認してください。")
            else:
                st.success(f"✅ {len(store_results)} 店舗の計算が完了しました。")
                for store_name, (rows, fmt) in store_results.items():
                    with st.expander(f"📌 {store_name} — {len(rows)} 名"):
                        df_preview = pd.DataFrame(rows)
                        st.dataframe(df_preview, use_container_width=True)

                excel_bytes = create_excel(store_results, period_str, std_hours, df_att=df_att)
                st.download_button(
                    label="📥 Excel ダウンロード",
                    data=excel_bytes,
                    file_name=f"給与計算_コメダ_{period_str}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                )
    else:
        st.info("上の2つのCSVをアップロードすると計算が開始できます。")


# ─── おはこタブ ─────────────────────────────
with tab_ohako:
    st.markdown("**📊 おはこは月集計CSVで計算します**（日次データなし → 早朝/越境/深夜有給は非対応）")
    col_a, col_b = st.columns(2)
    with col_a:
        st.subheader("📋 月集計CSV (Book7形式)")
        ohako_att_file = st.file_uploader(
            "月集計CSV（氏名/従業員番号/労働/総勤務/法定内/時間外/法定/深夜/欠勤/遅刻 列）",
            type=["csv"], key="ohako_att"
        )
    with col_b:
        st.subheader("👥 従業員エクスポート CSV")
        ohako_emp_file = st.file_uploader("employee_exports CSV", type=["csv"], key="ohako_emp")

    period_ohako = st.text_input("対象期間（出力ファイル名・タイトル用）", value="2026年5月", key="ohako_period")

    if ohako_att_file and ohako_emp_file:
        df_ohako = _read_csv_auto(ohako_att_file)
        df_emp2 = _read_csv_auto(ohako_emp_file)

        with st.expander("📊 データプレビュー"):
            st.write(f"**月集計データ** — {len(df_ohako)} 行 × {len(df_ohako.columns)} 列")
            st.dataframe(df_ohako.head(10), use_container_width=True)

        if st.button("🔄 給与計算を実行", type="primary", key="run_ohako"):
            with st.spinner("計算中..."):
                try:
                    store_results_o = process_ohako(df_ohako, df_emp2, std_hours)
                except Exception as e:
                    st.error(f"計算エラー: {e}")
                    st.exception(e)
                    st.stop()

            if not store_results_o:
                st.warning("計算結果が空です。")
            else:
                for store_name, (rows, fmt) in store_results_o.items():
                    with st.expander(f"📌 {store_name} — {len(rows)} 名", expanded=True):
                        st.dataframe(pd.DataFrame(rows), use_container_width=True)

                excel_bytes = create_excel(store_results_o, period_ohako, std_hours)
                st.download_button(
                    label="📥 Excel ダウンロード",
                    data=excel_bytes,
                    file_name=f"給与計算_おはこ_{period_ohako}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                )
    else:
        st.info("月集計CSVと従業員CSVをアップロードしてください。")

with st.expander("💡 CSVの作り方"):
    st.markdown("""
**コメダ**:
1. freeeのExcelを開く
2. Sheet1（日次勤怠一覧）→ 右クリック → 「移動またはコピー」→ 新しいブックへ → CSV(UTF-8)で保存
3. Sheet2（employee_exports）も同様にCSV保存
4. コメダタブで両方をアップロード

**おはこ**:
1. 月集計CSV（Book7.csv のような氏名/従業員番号/労働/総勤務/法定内/時間外/法定/深夜/欠勤/遅刻列を持つCSV）
2. 同じemployee_exports CSV
3. おはこタブで両方をアップロード
    """)
